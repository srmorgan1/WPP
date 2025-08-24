"""Tests for excel.py formatting utilities."""

import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from wpp.excel import format_all_excel_sheets, format_all_excel_sheets_comprehensive, format_excel_sheet, format_excel_sheet_comprehensive


@pytest.fixture
def sample_workbook():
    """Create a sample workbook with test data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add headers
    ws["A1"] = "Name"
    ws["B1"] = "Reference"
    ws["C1"] = "Error Message"
    ws["D1"] = "Amount"

    # Add data rows
    ws["A2"] = "John Doe"
    ws["B2"] = "REF123"
    ws["C2"] = "Invalid data format"
    ws["D2"] = 100.50

    ws["A3"] = "Jane Smith"
    ws["B3"] = "REF456"
    ws["C3"] = "Missing required field"
    ws["D3"] = 200.75

    return wb


@pytest.fixture
def empty_workbook():
    """Create an empty workbook."""
    wb = Workbook()
    wb.active
    return wb


@pytest.fixture
def single_row_workbook():
    """Create a workbook with only headers."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    return wb


class TestFormatExcelSheet:
    """Test the basic format_excel_sheet function."""

    def test_format_excel_sheet_basic(self, sample_workbook):
        """Test basic formatting with sample data."""
        ws = sample_workbook.active
        format_excel_sheet(ws, expand_columns=True)

        # Check header formatting
        assert ws["A1"].font.bold is True
        assert ws["A1"].alignment.horizontal == "center"
        assert ws["A1"].alignment.vertical == "center"

        # Check column widths are set
        assert ws.column_dimensions["A"].width > 0
        assert ws.column_dimensions["B"].width > 0
        assert ws.column_dimensions["C"].width > 0
        assert ws.column_dimensions["D"].width > 0

        # Check borders are applied
        assert ws["A1"].border.top.style == "medium"
        assert ws["A1"].border.left.style == "medium"
        assert ws["D1"].border.top.style == "medium"
        assert ws["D1"].border.right.style == "medium"

    def test_format_excel_sheet_expand_columns_false(self, sample_workbook):
        """Test formatting with expand_columns=False."""
        ws = sample_workbook.active
        format_excel_sheet(ws, expand_columns=False)

        # Should still format headers
        assert ws["A1"].font.bold is True
        assert ws["A1"].alignment.horizontal == "center"

        # Column widths should be constrained
        for col in ["A", "B", "C", "D"]:
            width = ws.column_dimensions[col].width
            assert width <= 50  # Should be constrained

    def test_format_excel_sheet_empty(self, empty_workbook):
        """Test formatting an empty worksheet."""
        ws = empty_workbook.active
        format_excel_sheet(ws)

        # Should not raise an error and handle empty sheet gracefully
        assert ws.max_row == 1
        assert ws.max_column == 1

    def test_format_excel_sheet_single_row(self, single_row_workbook):
        """Test formatting worksheet with only headers."""
        ws = single_row_workbook.active
        format_excel_sheet(ws)

        # Headers should be formatted
        assert ws["A1"].font.bold is True
        assert ws["B1"].font.bold is True

    def test_column_width_calculation(self, sample_workbook):
        """Test column width calculations for different column types."""
        ws = sample_workbook.active
        format_excel_sheet(ws, expand_columns=True)

        # Name column should have minimum width of 20
        name_width = ws.column_dimensions["A"].width
        assert name_width >= 20

        # Reference column should have minimum width of 15
        ref_width = ws.column_dimensions["B"].width
        assert ref_width >= 15

        # Error column should have minimum width of 30
        error_width = ws.column_dimensions["C"].width
        assert error_width >= 30


class TestFormatExcelSheetComprehensive:
    """Test the comprehensive format_excel_sheet_comprehensive function."""

    def test_comprehensive_formatting_basic(self, sample_workbook):
        """Test comprehensive formatting with sample data."""
        ws = sample_workbook.active
        format_excel_sheet_comprehensive(ws, "TestSheet", expand_columns=True)

        # Check header formatting - should have white font and blue background
        assert ws["A1"].font.color.rgb == "00FFFFFF"
        assert ws["A1"].fill.start_color.rgb == "00366092"
        assert ws["A1"].alignment.horizontal == "center"

        # Check alternating row colors
        # Row 2 (even) should be light grey
        assert ws["A2"].fill.start_color.rgb == "00F5F5F5"
        # Row 3 (odd) should be white
        assert ws["A3"].fill.start_color.rgb == "00FFFFFF"

        # Error column cells should have red font
        assert ws["C2"].font.color.rgb == "00D32F2F"
        assert ws["C3"].font.color.rgb == "00D32F2F"

    def test_comprehensive_expand_columns_false(self, sample_workbook):
        """Test comprehensive formatting with expand_columns=False."""
        ws = sample_workbook.active
        format_excel_sheet_comprehensive(ws, expand_columns=False)

        # Error columns should be capped at 60
        error_width = ws.column_dimensions["C"].width
        assert error_width <= 60

        # Reference columns should be capped at 25
        ref_width = ws.column_dimensions["B"].width
        assert ref_width <= 25

    def test_comprehensive_empty_sheet(self, empty_workbook):
        """Test comprehensive formatting on empty sheet."""
        ws = empty_workbook.active
        format_excel_sheet_comprehensive(ws, "EmptySheet")

        # Should handle empty sheet without errors
        assert ws.max_row == 1

    def test_comprehensive_borders(self, sample_workbook):
        """Test that comprehensive formatting applies borders correctly."""
        ws = sample_workbook.active
        format_excel_sheet_comprehensive(ws)

        # Check outer borders
        assert ws["A1"].border.left.style == "medium"
        assert ws["A1"].border.top.style == "medium"
        assert ws["D3"].border.right.style == "medium"
        assert ws["D3"].border.bottom.style == "medium"


class TestFormatAllExcelSheets:
    """Test functions that format all sheets in an Excel writer."""

    def test_format_all_excel_sheets(self):
        """Test formatting all sheets with basic formatting."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            tmp_path = Path(tmp_file.name)

        try:
            # Create test data
            df1 = pd.DataFrame({"Name": ["John", "Jane"], "Value": [100, 200]})
            df2 = pd.DataFrame({"Reference": ["REF1", "REF2"], "Error Message": ["Error1", "Error2"]})

            # Write to Excel with multiple sheets
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                df1.to_excel(writer, sheet_name="Sheet1", index=False)
                df2.to_excel(writer, sheet_name="Sheet2", index=False)

                # Apply formatting
                format_all_excel_sheets(writer)

                # Check that formatting was applied to both sheets
                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    # Headers should be bold
                    for cell in ws[1]:
                        if cell.value:
                            assert cell.font.bold is True

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_format_all_excel_sheets_comprehensive(self):
        """Test formatting all sheets with comprehensive formatting."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            tmp_path = Path(tmp_file.name)

        try:
            # Create test data
            df = pd.DataFrame({"Name": ["Test"], "Error Message": ["Test error"]})

            # Write to Excel
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="TestSheet", index=False)

                # Apply comprehensive formatting
                format_all_excel_sheets_comprehensive(writer)

                # Check comprehensive formatting was applied
                ws = writer.sheets["TestSheet"]
                # Headers should have blue background
                assert ws["A1"].fill.start_color.rgb == "00366092"
                # Data rows should have alternating colors
                assert ws["A2"].fill.start_color.rgb == "00F5F5F5"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()


class TestEdgeCases:
    """Test edge cases and error conditions."""

    def test_worksheet_with_none_values(self):
        """Test formatting worksheet with None/null values."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A2"] = None
        ws["A3"] = ""
        ws["A4"] = "Value"

        # Should not raise an error
        format_excel_sheet(ws)
        format_excel_sheet_comprehensive(ws)

        # Headers should still be formatted
        assert ws["A1"].font.bold is True

    def test_worksheet_with_very_long_content(self):
        """Test formatting with very long cell content."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Error Message"
        ws["A2"] = "This is a very long error message " * 10  # Very long content

        format_excel_sheet(ws, expand_columns=False)

        # Column width should be capped when not expanding
        width = ws.column_dimensions["A"].width
        assert width <= 50

    def test_worksheet_with_numeric_headers(self):
        """Test formatting with numeric header values."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 123
        ws["B1"] = 456.789
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"

        # Should not raise an error
        format_excel_sheet_comprehensive(ws)

        # Numeric headers should still be formatted
        assert ws["A1"].font.color.rgb == "00FFFFFF"
        assert ws["B1"].font.color.rgb == "00FFFFFF"

    def test_single_cell_worksheet(self):
        """Test formatting worksheet with single cell."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Single"

        format_excel_sheet(ws)
        format_excel_sheet_comprehensive(ws)

        # Single cell should be formatted as header
        assert ws["A1"].font.bold is True
        assert ws["A1"].border.top.style == "medium"
        assert ws["A1"].border.bottom.style == "medium"
        assert ws["A1"].border.left.style == "medium"
        assert ws["A1"].border.right.style == "medium"

    def test_column_width_edge_cases(self):
        """Test column width calculations for edge cases."""
        wb = Workbook()
        ws = wb.active

        # Test with empty string
        ws["A1"] = ""
        ws["B1"] = "   "  # Whitespace only
        ws["C1"] = "Normal Header"

        format_excel_sheet(ws, expand_columns=True)

        # All columns should have minimum widths
        assert ws.column_dimensions["A"].width >= 10
        assert ws.column_dimensions["B"].width >= 10
        assert ws.column_dimensions["C"].width >= 10


class TestSpecificColumnTypes:
    """Test formatting behavior for specific column types."""

    def test_error_column_formatting(self):
        """Test specific formatting for Error columns."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Error Details"
        ws["A2"] = "Critical error occurred"

        format_excel_sheet_comprehensive(ws, expand_columns=True)

        # Error column should have minimum width of 30
        width = ws.column_dimensions["A"].width
        assert width >= 30

        # Error text should be red
        assert ws["A2"].font.color.rgb == "00D32F2F"

    def test_reference_column_formatting(self):
        """Test specific formatting for Reference columns."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Reference Number"
        ws["A2"] = "REF123456"

        format_excel_sheet(ws, expand_columns=True)

        # Reference column should have minimum width of 15
        width = ws.column_dimensions["A"].width
        assert width >= 15

    def test_name_column_formatting(self):
        """Test specific formatting for Name columns."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Customer Name"
        ws["A2"] = "John Doe Smith"

        format_excel_sheet(ws, expand_columns=True)

        # Name column should have minimum width of 20
        width = ws.column_dimensions["A"].width
        assert width >= 20
