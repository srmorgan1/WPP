"""Excel formatting utilities for WPP system."""

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def format_excel_sheet(worksheet, expand_columns: bool = True) -> None:
    """Format Excel sheet with headers and outer border."""

    # Define styles
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Auto-size columns with content-aware sizing (matching comprehensive function)
    for column in worksheet.columns:
        column_letter = column[0].column_letter
        max_length = 0

        for cell in column:
            try:
                if cell.value is not None:
                    # Convert to string and measure actual display length
                    cell_text = str(cell.value).strip()
                    cell_length = len(cell_text)
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass

        # Content-aware width calculation (matching comprehensive function)
        if expand_columns:
            # Expand to fit all content with reasonable minimums and generous padding
            if column[0].value and "Error" in str(column[0].value):
                adjusted_width = max(max_length + 5, 30)  # Error columns: min 30, +5 padding
            elif column[0].value and "Reference" in str(column[0].value):
                adjusted_width = max(max_length + 3, 15)  # Reference columns: min 15, +3 padding
            elif column[0].value and "Name" in str(column[0].value):
                adjusted_width = max(max_length + 3, 20)  # Name columns: min 20, +3 padding
            else:
                adjusted_width = max(max_length + 3, 10)  # Other columns: min 10, +3 padding
        else:
            # Use original simple limits
            if max_length == 0:
                adjusted_width = 10
            elif max_length < 8:
                adjusted_width = 10
            elif max_length > 100:
                adjusted_width = 50
            else:
                adjusted_width = max_length + 3

        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Apply medium outer border to the entire table (matching comprehensive function)
    if worksheet.max_row > 0 and worksheet.max_column > 0:
        min_col = 1
        max_col = worksheet.max_column
        min_row = 1
        max_row = worksheet.max_row

        # Apply outer border to the entire table
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row=row, column=col)

                # Determine which borders to apply
                border_left = Side(style="medium") if col == min_col else None
                border_right = Side(style="medium") if col == max_col else None
                border_top = Side(style="medium") if row == min_row else None
                border_bottom = Side(style="medium") if row == max_row else None

                cell.border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom)

    # Format header row (row 1)
    if worksheet.max_row > 0:
        for cell in worksheet[1]:  # First row
            cell.font = header_font
            cell.alignment = center_alignment


def format_excel_sheet_comprehensive(worksheet, sheet_name: str = "", expand_columns: bool = True) -> None:
    """Comprehensive Excel sheet formatting with conditional styling."""

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")  # White text
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Blue background
    error_font = Font(color="D32F2F")  # Red text for errors

    # Bold outer border
    # bold_border = Border(left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium"))

    # Alternating row fills
    light_grey_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Light grey
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

    # Auto-size columns with content-aware sizing
    for column in worksheet.columns:
        column_letter = column[0].column_letter
        max_length = 0

        for cell in column:
            try:
                if cell.value is not None:
                    # Convert to string and measure actual display length
                    cell_text = str(cell.value).strip()
                    cell_length = len(cell_text)

                    # For non-expanding mode, cap error columns at 60 characters
                    if not expand_columns and column[0].value and "Error" in str(column[0].value):
                        cell_length = min(cell_length, 60)

                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass

        # Content-aware width calculation
        if expand_columns:
            # Expand to fit all content with reasonable minimums and generous padding
            if column[0].value and "Error" in str(column[0].value):
                adjusted_width = max(max_length + 5, 30)  # Error columns: min 30, +5 padding
            elif column[0].value and "Reference" in str(column[0].value):
                adjusted_width = max(max_length + 3, 15)  # Reference columns: min 15, +3 padding
            elif column[0].value and "Name" in str(column[0].value):
                adjusted_width = max(max_length + 3, 20)  # Name columns: min 20, +3 padding
            else:
                adjusted_width = max(max_length + 3, 10)  # Other columns: min 10, +3 padding
        else:
            # Use original limits
            if column[0].value and "Error" in str(column[0].value):
                adjusted_width = min(max(max_length + 2, 30), 60)  # Error columns: min 30, max 60
            elif column[0].value and "Reference" in str(column[0].value):
                adjusted_width = min(max(max_length + 2, 15), 25)  # Reference columns: min 15, max 25
            elif column[0].value and "Name" in str(column[0].value):
                adjusted_width = min(max(max_length + 2, 20), 40)  # Name columns: min 20, max 40
            else:
                adjusted_width = min(max(max_length + 2, 10), 30)  # Other columns: min 10, max 30

        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Determine table range for outer border
    if worksheet.max_row > 0 and worksheet.max_column > 0:
        min_col = 1
        max_col = worksheet.max_column
        min_row = 1
        max_row = worksheet.max_row

        # Apply outer border to the entire table
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row=row, column=col)

                # Determine which borders to apply
                border_left = Side(style="medium") if col == min_col else None
                border_right = Side(style="medium") if col == max_col else None
                border_top = Side(style="medium") if row == min_row else None
                border_bottom = Side(style="medium") if row == max_row else None

                cell.border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom)

    # Format header row
    if worksheet.max_row > 0:
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Format data rows with alternating colors
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
        for col_num, cell in enumerate(row, 1):
            # Highlight error text in red
            if worksheet.cell(1, col_num).value and "Error" in str(worksheet.cell(1, col_num).value):
                cell.font = error_font

            # Alternate row colors: light grey and white
            if row_num % 2 == 0:
                cell.fill = light_grey_fill  # Even rows: light grey
            else:
                cell.fill = white_fill  # Odd rows: white


def format_all_excel_sheets(excel_writer: pd.ExcelWriter) -> None:
    """Format all sheets in the Excel writer using the basic formatting."""
    for sheet_name in excel_writer.sheets:
        worksheet = excel_writer.sheets[sheet_name]
        format_excel_sheet(worksheet, expand_columns=True)


def format_all_excel_sheets_comprehensive(excel_writer: pd.ExcelWriter) -> None:
    """Format all sheets in the Excel writer using comprehensive formatting."""
    for sheet_name in excel_writer.sheets:
        worksheet = excel_writer.sheets[sheet_name]
        format_excel_sheet_comprehensive(worksheet, sheet_name, expand_columns=True)
