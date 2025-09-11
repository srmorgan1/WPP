"""Excel/CSV data loading functions for WPP management system."""

import logging
import re
import sqlite3

import pandas as pd
from dateutil import parser
from openpyxl import load_workbook

from ..data_classes import ChargeData
from ..database.database_commands import DatabaseCommandExecutor, InsertChargeCommand, InsertTenantCommand, UpdateTenantNameCommand
from ..database.db import get_last_insert_id, get_single_value
from ..output.output_handler import OutputHandler
from ..ref_matcher import getPropertyBlockAndTenantRefs as getPropertyBlockAndTenantRefs_strategy
from ..utils.exceptions import database_transaction

# Set up module logger
logger = logging.getLogger(__name__)

#
# Constants
#
CLIENT_CREDIT_ACCOUNT_NUMBER = "06000792"

#
# SQL
#
INSERT_PROPERTY_SQL = "INSERT INTO Properties (property_ref, property_name) VALUES (?, Null);"
INSERT_BLOCK_SQL = "INSERT INTO Blocks (block_ref, block_name, type, property_id) VALUES (?, Null, ?, ?);"
INSERT_BLOCK_SQL2 = "INSERT INTO Blocks (block_ref, block_name, type, property_id) VALUES (?, ?, ?, ?);"
INSERT_TENANT_SQL = "INSERT INTO Tenants (tenant_ref, tenant_name, block_id) VALUES (?, ?, ?);"
INSERT_BANK_ACCOUNT_SQL = "INSERT INTO Accounts (sort_code, account_number, account_type, property_or_block, client_ref, account_name, block_id) VALUES (?, ?, ?, ?, ?, ?, ?);"
INSERT_BANK_ACCOUNT_BALANCE_SQL = "INSERT INTO AccountBalances (current_balance, available_balance, at_date, account_id) VALUES (?, ?, ?, ?);"
INSERT_KEY_TABLE_SQL = "INSERT INTO Key_{} (value) VALUES (?);"
INSERT_IRREGULAR_TRANSACTION_REF_SQL = "INSERT INTO IrregularTransactionRefs (tenant_ref, transaction_ref_pattern) VALUES (?, ?);"
INSERT_CHARGES_SQL = "INSERT INTO Charges (fund_id, category_id, type_id, at_date, amount, block_id) VALUES (?, ?, ?, ?, ?, ?);"

SELECT_TENANT_ID_SQL = "SELECT tenant_id FROM Tenants WHERE tenant_ref = ?;"
SELECT_ID_FROM_REF_SQL = "SELECT ID FROM {} WHERE {}_ref = '{}';"
SELECT_ID_FROM_KEY_TABLE_SQL = "SELECT ID FROM Key_{} WHERE value = ?;"
SELECT_PROPERTY_ID_FROM_REF_SQL = "SELECT ID FROM Properties WHERE property_ref = ? AND property_name IS NULL;"
SELECT_BANK_ACCOUNT_SQL = "SELECT ID FROM Blocks WHERE ID = ? AND account_number IS Null;"
SELECT_BANK_ACCOUNT_SQL1 = "SELECT ID FROM Accounts WHERE sort_code = ? AND account_number = ?;"
SELECT_BANK_ACCOUNT_BALANCE_SQL = "SELECT ID FROM AccountBalances WHERE at_date = ? AND account_id = ?;"
SELECT_BLOCK_NAME_SQL = "SELECT block_name FROM Blocks WHERE block_ref = ?;"
SELECT_TENANT_NAME_BY_ID_SQL = "SELECT tenant_name FROM Tenants WHERE ID = ?;"
SELECT_IRREGULAR_TRANSACTION_REF_ID_SQL = "select ID from IrregularTransactionRefs where tenant_ref = ? and transaction_ref_pattern = ?;"
SELECT_CHARGES_SQL = "SELECT ID FROM Charges WHERE fund_id = ? AND category_id = ? and type_id = ? and block_id = ? and at_date = ?;"

UPDATE_BLOCK_ACCOUNT_NUMBER_SQL = "UPDATE Blocks SET account_number = ? WHERE ID = ? AND account_number IS Null;"
UPDATE_PROPERTY_DETAILS_SQL = "UPDATE Properties SET property_name = ? WHERE ID = ?;"
UPDATE_BLOCK_NAME_SQL = "UPDATE Blocks SET block_name = ? WHERE ID = ?;"
UPDATE_TENANT_NAME_SQL = "UPDATE Tenants SET tenant_name = ? WHERE ID = ?;"

# Charge types
AUTH_CREDITORS = "Auth Creditors"
AVAILABLE_FUNDS = "Available Funds"
SC_FUND = "SC Fund"

#
# Helper functions
#


def getPropertyBlockAndTenantRefs(reference: str) -> tuple[str | None, str | None, str | None]:
    """Wrapper for the strategy pattern function import."""
    return getPropertyBlockAndTenantRefs_strategy(reference)


def get_id(db_cursor: sqlite3.Cursor, sql: str, args_tuple: tuple = ()) -> int | None:
    return get_single_value(db_cursor, sql, args_tuple)


def get_id_from_ref(db_cursor: sqlite3.Cursor, table_name: str, field_name: str, ref_name: str) -> int | None:
    sql = SELECT_ID_FROM_REF_SQL.format(table_name, field_name, ref_name)
    db_cursor.execute(sql)
    _id = db_cursor.fetchone()
    if _id:
        return _id[0]
    else:
        return None


def get_id_from_key_table(db_cursor: sqlite3.Cursor, key_table_name: str, value: str) -> int:
    sql = SELECT_ID_FROM_KEY_TABLE_SQL.format(key_table_name)
    db_cursor.execute(sql, (value,))
    _id = db_cursor.fetchone()
    if _id:
        return _id[0]
    else:
        sql = INSERT_KEY_TABLE_SQL.format(key_table_name)
        db_cursor.execute(sql, (value,))
        return get_last_insert_id(db_cursor, f"Key_{key_table_name}")


def _validate_reference_parsing(references_df: pd.DataFrame, file_path: str, reference_column: str, name_column: str, file_type: str) -> list[dict]:
    """
    Validate that references in a dataframe can be parsed correctly.

    Args:
        references_df: DataFrame containing reference data
        file_path: File path for error messages
        reference_column: Name of the reference column
        name_column: Name of the name column (for context)
        file_type: Type of file for error messages ('Tenants', 'Estates', 'General Idents')

    Returns:
        List of validation issues found
    """
    validation_issues = []

    for index, row in references_df.iterrows():
        reference = row[reference_column]
        name = row[name_column] if name_column in row and pd.notna(row[name_column]) else ""

        # Skip blank/empty references
        if pd.isna(reference) or str(reference).strip() == "":
            continue

        # Try to parse the reference
        property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefs(str(reference).strip())

        # Check if parsing was successful based on file type
        parsing_failed = False
        error_msg = ""

        if file_type == "Tenants":
            # For tenants file, we need all three parts (property, block, tenant)
            if not all([property_ref, block_ref, tenant_ref]):
                parsing_failed = True
                error_msg = f"Tenant reference '{reference}' could not be parsed into property-block-tenant format"
        elif file_type == "Estates":
            # For estates file, we need at least property reference
            if not property_ref:
                parsing_failed = True
                error_msg = f"Estate reference '{reference}' could not be parsed to extract property reference"
        elif file_type == "General Idents":
            # For general idents file, we need tenant reference to be parseable
            if not tenant_ref:
                parsing_failed = True
                error_msg = f"Tenant reference '{reference}' could not be parsed to extract tenant reference"

        if parsing_failed:
            validation_issues.append(
                {
                    "Row Number": index + 2,  # +2 because pandas is 0-indexed and Excel has header row
                    "Reference": reference,
                    "Name": name,
                    "Property Ref": property_ref or "N/A",
                    "Block Ref": block_ref or "N/A",
                    "Tenant Ref": tenant_ref or "N/A",
                    "Error": error_msg,
                }
            )

    return validation_issues


def _report_reference_parsing_errors(validation_issues: list[dict], file_path: str, file_type: str, output_handler: OutputHandler) -> None:
    """
    Report reference parsing errors to log and output handler.

    Args:
        validation_issues: List of error dictionaries
        file_path: File path for error messages
        file_type: Type of file for sheet naming
        output_handler: OutputHandler for outputting errors
    """
    if not validation_issues:
        return

    logger.error(f"Found {len(validation_issues)} reference parsing issues in {file_path}")
    for error in validation_issues:
        logger.error(f"Row {error['Row Number']}: {error['Error']} - Name: {error['Name']}")

    # Write errors to output handler with file-type specific sheet name
    sheet_name = f"{file_type} Import Problems"
    errors_df = pd.DataFrame(validation_issues)
    output_handler.add_sheet(sheet_name, errors_df, {"file_path": file_path, "error_count": len(validation_issues)})


def _validate_account_uniqueness(bank_accounts_df: pd.DataFrame, bank_accounts_file: str, output_handler: OutputHandler) -> None:
    """
    Validate that each CL account appears only once per block.

    Args:
        bank_accounts_df: DataFrame containing account data
        bank_accounts_file: File path for error messages
        output_handler: OutputHandler for outputting validation issues

    Raises:
        ValueError: If duplicate CL accounts are found for the same block
    """
    # Group by Reference and Account Type, count occurrences
    cl_accounts = bank_accounts_df[bank_accounts_df["Account Type"] == "CL"]
    duplicates = cl_accounts.groupby(["Reference", "Account Type"]).size()
    duplicate_cl_accounts = duplicates[duplicates > 1]

    if len(duplicate_cl_accounts) > 0:
        logger.error(f"Found {len(duplicate_cl_accounts)} duplicate CL account(s) in {bank_accounts_file}")

        # Create detailed error report
        duplicate_details = []
        for (reference, account_type), count in duplicate_cl_accounts.items():
            matching_rows = cl_accounts[cl_accounts["Reference"] == reference]
            for _, row in matching_rows.iterrows():
                duplicate_details.append(
                    {
                        "Reference": reference,
                        "Account Type": account_type,
                        "Sort Code": row["Sort Code"],
                        "Account Number": row["Account Number"],
                        "Account Name": row["Account Name"],
                        "Client Reference": row["Client Reference"],
                        "Property Or Block": row["Property Or Block"],
                        "Error": f"Duplicate CL account found {count} times for reference {reference}",
                    }
                )

        # Report to output handler
        errors_df = pd.DataFrame(duplicate_details)
        output_handler.add_sheet("Bank Account Duplicate Problems", errors_df, {"file_path": bank_accounts_file, "error_count": len(duplicate_details)}, is_critical=True)

        raise ValueError(f"Found {len(duplicate_cl_accounts)} duplicate CL account(s) - import halted")


def _validate_account_designation_consistency(bank_accounts_df: pd.DataFrame, bank_accounts_file: str, output_handler: OutputHandler) -> list[dict]:
    """
    Validate that manual Property/Block designation matches the reference format for all accounts.

    Args:
        bank_accounts_df: DataFrame containing account data
        bank_accounts_file: File path for error messages
        output_handler: OutputHandler for outputting validation issues

    Returns:
        List of validation issues found
    """
    validation_issues = []

    for index, row in bank_accounts_df.iterrows():
        reference = row["Reference"]
        property_or_block = row["Property Or Block"]
        sort_code = row["Sort Code"]
        account_number = row["Account Number"]
        account_name = row["Account Name"]
        client_ref = row["Client Reference"]

        # Skip accounts with blank/empty references
        if pd.isna(reference) or str(reference).strip() == "":
            continue

        # Parse the reference to determine what it should be
        property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefs(str(reference).strip())

        # Determine expected designation based on reference format
        expected_designation = None
        if property_ref and block_ref and not tenant_ref:
            # Property-Block format (e.g., "034-01") should be Block
            expected_designation = "B"
        elif property_ref and not block_ref and not tenant_ref:
            # Property-only format (e.g., "034") should be Property
            expected_designation = "P"

        # Normalize current designation
        current_designation = None
        if property_or_block and str(property_or_block).strip():
            prop_or_block_upper = str(property_or_block).strip().upper()
            if prop_or_block_upper in ["PROPERTY", "P"]:
                current_designation = "P"
            elif prop_or_block_upper in ["BLOCK", "B"]:
                current_designation = "B"

        # Check for inconsistency
        if expected_designation and current_designation and expected_designation != current_designation:
            designation_names = {"P": "Property", "B": "Block"}
            validation_issues.append(
                {
                    "Row Number": index + 2,  # +2 because pandas is 0-indexed and Excel has header row
                    "Reference": reference,
                    "Account Name": account_name,
                    "Sort Code": sort_code,
                    "Account Number": account_number,
                    "Client Reference": client_ref,
                    "Current Designation": property_or_block,
                    "Expected Designation": designation_names.get(expected_designation, expected_designation),
                    "Property Ref": property_ref or "N/A",
                    "Block Ref": block_ref or "N/A",
                    "Error": f"Reference format suggests {designation_names.get(expected_designation)} but marked as {property_or_block}",
                }
            )

    if validation_issues:
        logger.warning(f"Found {len(validation_issues)} Property/Block designation issues in {bank_accounts_file}")
        errors_df = pd.DataFrame(validation_issues)
        output_handler.add_sheet("Bank Account Designation Problems", errors_df, {"file_path": bank_accounts_file, "error_count": len(validation_issues)})

    return validation_issues


def _read_properties_df(properties_xls_file: str) -> pd.DataFrame:
    """Read properties Excel file into DataFrame with proper data types."""
    properties_df = pd.read_excel(properties_xls_file, dtype={"Reference": str})
    properties_df.fillna("", inplace=True)
    return properties_df


def _is_valid_reference(reference: str) -> bool:
    """Check if a reference is valid for processing."""
    return reference and len(reference.strip()) > 0


def _process_property(csr: sqlite3.Cursor, property_ref: str) -> tuple[int, int]:
    """Process a property reference, creating it if needed. Returns (property_id, properties_added)."""
    property_id = get_id_from_ref(csr, "Properties", "property", property_ref)
    if not property_id:
        csr.execute(INSERT_PROPERTY_SQL, (property_ref,))
        property_id = get_id_from_ref(csr, "Properties", "property", property_ref)
        logger.debug(f"\tAdding property {property_ref}")
        return property_id, 1
    return property_id, 0


def _process_block(csr: sqlite3.Cursor, block_ref: str, property_id: int) -> tuple[int, int]:
    """Process a block reference, creating it if needed. Returns (block_id, blocks_added)."""
    block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
    if not block_id:
        block_type = "P" if block_ref and block_ref.endswith("00") else "B"
        csr.execute(INSERT_BLOCK_SQL, (block_ref, block_type, property_id))
        block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
        logger.debug(f"\tAdding block {block_ref}")
        return block_id, 1
    return block_id, 0


def _process_tenant(csr: sqlite3.Cursor, tenant_ref: str, tenant_name: str, block_id: int) -> int:
    """Process a tenant reference, creating or updating it if needed. Returns tenants_added."""
    executor = DatabaseCommandExecutor(csr, logger)
    tenant_id = get_id_from_ref(csr, "Tenants", "tenant", tenant_ref)

    if not tenant_id:
        # Create new tenant
        command = InsertTenantCommand(tenant_ref, tenant_name, block_id, INSERT_TENANT_SQL)
        executor.execute(command)
        logger.debug(f"\tAdding tenant {tenant_ref}")
        return 1
    else:
        # Check if we need to update the tenant name
        old_tenant_name = get_single_value(csr, SELECT_TENANT_NAME_BY_ID_SQL, (tenant_id,))
        if old_tenant_name != tenant_name and tenant_name.strip():
            command = UpdateTenantNameCommand(tenant_id, tenant_name, UPDATE_TENANT_NAME_SQL)
            executor.execute(command)
            logger.debug(f"\tUpdating tenant name for {tenant_ref} from '{old_tenant_name}' to '{tenant_name}'")
        return 0


def calculateSCFund(auth_creditors: float, available_funds: float, property_ref: str, block_ref: str) -> float:
    """Calculate Service Charge Fund value based on business rules."""
    # TODO: This should be encoded in a user-supplied rules spreadsheet for generality
    if property_ref == "035":
        return available_funds
    else:
        return auth_creditors + available_funds


def _validate_qube_spreadsheet(workbook_sheet) -> None:
    """Validate that the spreadsheet is a valid Qube balances report."""
    A1_cell_value = workbook_sheet.cell(1, 1).value
    B1_cell_value = workbook_sheet.cell(1, 2).value
    produced_date_cell_value = workbook_sheet.cell(3, 1).value

    if not isinstance(produced_date_cell_value, str):
        raise ValueError(f"The produced date cell value is not a string: {produced_date_cell_value}")

    cell_values_actual = [workbook_sheet.cell(5, i + 1).value for i in range(0, 5)]
    cell_values_expected = [
        "Property / Fund",
        "Bank",
        "Excluded VAT",
        AUTH_CREDITORS,
        AVAILABLE_FUNDS,
    ]

    if not all(actual == expected for actual, expected in zip(cell_values_actual, cell_values_expected)):
        raise ValueError(f"Unexpected Qube spreadsheet format. Expected headers: {cell_values_expected}, got: {cell_values_actual}")


def _extract_qube_date(workbook_sheet) -> str:
    """Extract the date from Qube spreadsheet."""
    produced_date_cell_value = workbook_sheet.cell(3, 1).value
    # Parse the date from the "Produced on DD/MM/YYYY at HH:MM" format
    date_match = re.search(r"(\d{2}/\d{2}/\d{4})", produced_date_cell_value)
    if date_match:
        return parser.parse(date_match.group(1), dayfirst=True).strftime("%Y-%m-%d")
    else:
        raise ValueError(f"Could not extract date from: {produced_date_cell_value}")


def _prepare_qube_dataframe(qube_eod_balances_xls_file: str) -> pd.DataFrame:
    """Prepare the Qube dataframe for processing."""
    # Read from row 5 (0-indexed row 4) to skip headers
    qube_eod_balances_df = pd.read_excel(qube_eod_balances_xls_file, skiprows=4, dtype={"Property / Fund": str})
    qube_eod_balances_df.fillna("", inplace=True)

    # Rename columns for consistency
    column_mapping = {"Property / Fund": "PropertyCode / Fund", "Bank": "PropertyName / Category"}
    qube_eod_balances_df = qube_eod_balances_df.rename(columns=column_mapping)

    return qube_eod_balances_df


def _process_qube_data(
    csr, property_code_or_fund: str, property_ref: str, block_ref: str, block_name: str, fund: str, category: str, auth_creditors: float, available_funds: float, at_date: str, type_ids: dict
) -> tuple[int, dict | None]:
    """
    Process Qube fund/category data and add charges to database.

    Returns:
        Tuple of (charges_added, error_info)
    """
    # Ensure block exists in database
    block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
    if not block_id:
        # Try to create the block
        property_id = get_id_from_ref(csr, "Properties", "property", property_ref)
        if not property_id:
            # Cannot proceed without property
            error_msg = f"Property '{property_ref}' does not exist in database"
            logger.error(f"Cannot process Qube balances for block '{block_ref}'. {error_msg}")
            error_details = {"Property Reference": property_ref, "Block Reference": block_ref, "Block Name": block_name, "Fund": fund, "Category": category, "Error": error_msg}
            return 0, error_details

        # Create the block
        block_type = "P" if block_ref and block_ref.endswith("00") else "B"
        csr.execute(INSERT_BLOCK_SQL2, (block_ref, block_name, block_type, property_id))
        block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
        logger.debug(f"\tCreated block {block_ref} for Qube processing")

    # Update block name if it's empty
    existing_block_name = get_single_value(csr, SELECT_BLOCK_NAME_SQL, (block_ref,))
    if not existing_block_name and block_name:
        csr.execute(UPDATE_BLOCK_NAME_SQL, (block_name, block_id))
        logger.debug(f"\tUpdated block name for {block_ref} to '{block_name}'")

    # Get fund and category IDs
    fund_id = get_id_from_key_table(csr, "fund", fund)
    category_id = get_id_from_key_table(csr, "category", category)

    # Add charges
    charges_added = 0
    executor = DatabaseCommandExecutor(csr, logger)

    # Available funds charge
    available_funds_charge = ChargeData(fund_id, category_id, type_ids["available_funds"], at_date, available_funds, block_id)
    charges_id = get_id(csr, SELECT_CHARGES_SQL, (fund_id, category_id, type_ids["available_funds"], block_id, at_date))
    if not charges_id:
        command = InsertChargeCommand(available_funds_charge, INSERT_CHARGES_SQL)
        executor.execute(command)
        charges_added += 1

    # Add auth creditors and SC fund charges for Service Charge and Tenant Recharge
    if property_code_or_fund in ["Service Charge", "Tenant Recharge"]:
        # Auth creditors charge
        auth_creditors_charge = ChargeData(fund_id, category_id, type_ids["auth_creditors"], at_date, auth_creditors, block_id)
        charges_id = get_id(csr, SELECT_CHARGES_SQL, (fund_id, category_id, type_ids["auth_creditors"], block_id, at_date))
        if not charges_id:
            command = InsertChargeCommand(auth_creditors_charge, INSERT_CHARGES_SQL)
            executor.execute(command)
            charges_added += 1

        # SC Fund charge
        sc_fund_val = calculateSCFund(auth_creditors, available_funds, property_ref, block_ref)
        sc_fund_charge = ChargeData(fund_id, category_id, type_ids["sc_fund"], at_date, sc_fund_val, block_id)
        charges_id = get_id(csr, SELECT_CHARGES_SQL, (fund_id, category_id, type_ids["sc_fund"], block_id, at_date))
        if not charges_id:
            command = InsertChargeCommand(sc_fund_charge, INSERT_CHARGES_SQL)
            executor.execute(command)
            charges_added += 1

    return charges_added, None


def _report_qube_import_errors(qube_import_errors: list[dict], qube_file: str, output_handler: OutputHandler) -> None:
    """
    Report Qube import errors to log and output handler.

    Args:
        qube_import_errors: List of error dictionaries
        qube_file: File path for error messages
        output_handler: OutputHandler for outputting errors
    """
    if not qube_import_errors:
        return

    logger.error(f"Found {len(qube_import_errors)} Qube import issues in {qube_file}")
    for error in qube_import_errors:
        logger.error(f"Block {error['Block Reference']}: {error['Error']}")

    # Write errors to output handler
    errors_df = pd.DataFrame(qube_import_errors)
    output_handler.add_sheet("Qube Import Problems", errors_df, {"file_path": qube_file, "error_count": len(qube_import_errors)}, is_critical=True)


#
# Main import functions
#


def importPropertiesFile(db_conn: sqlite3.Connection, properties_xls_file: str, output_handler: OutputHandler | None = None) -> bool:
    """Import properties data from Excel file into database."""
    properties_df = _read_properties_df(properties_xls_file)

    # Validate tenant reference parsing before import
    if output_handler is not None:
        validation_issues = _validate_reference_parsing(properties_df, properties_xls_file, "Reference", "Name", "Tenants")
        _report_reference_parsing_errors(validation_issues, properties_xls_file, "Tenants", output_handler)

    num_properties_added_to_db = 0
    num_blocks_added_to_db = 0
    num_tenants_added_to_db = 0
    current_data = {}

    with database_transaction(db_conn, logger, "importing properties file") as csr:
        for index, row in properties_df.iterrows():
            reference = row["Reference"]
            tenant_name = row["Name"]
            current_data = {"reference": reference, "tenant_name": tenant_name}

            if not _is_valid_reference(reference):
                continue

            property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefs(reference)
            if not all((property_ref, block_ref, tenant_ref)):
                logger.warning(f"\tUnable to parse tenant reference {reference}, will not add to the database.")
                continue

            current_data.update({"property_ref": property_ref, "block_ref": block_ref, "tenant_ref": tenant_ref})

            # Type assertions for mypy since we know they're not None after the all() check
            assert property_ref is not None
            assert block_ref is not None
            assert tenant_ref is not None

            try:
                prop_id, props_added = _process_property(csr, property_ref)
                num_properties_added_to_db += props_added

                blk_id, blocks_added = _process_block(csr, block_ref, prop_id)
                num_blocks_added_to_db += blocks_added

                tenants_added = _process_tenant(csr, tenant_ref, tenant_name, blk_id)
                num_tenants_added_to_db += tenants_added
            except Exception:
                logger.error(f"Failed to process record: {current_data}")
                raise

    logger.info(f"{num_properties_added_to_db} properties added to the database.")
    logger.info(f"{num_blocks_added_to_db} blocks added to the database.")
    logger.info(f"{num_tenants_added_to_db} tenants added to the database.")

    # Tenants validation errors don't stop processing
    return False


def importEstatesFile(db_conn: sqlite3.Connection, estates_xls_file: str, output_handler: OutputHandler | None = None) -> bool:
    """Import estates data from Excel file into database."""
    # Read Excel spreadsheet into dataframe
    estates_df = pd.read_excel(estates_xls_file, dtype={"Reference": str})
    estates_df.fillna("", inplace=True)

    # Validate estate reference parsing before import
    has_critical_errors = False
    if output_handler is not None:
        validation_issues = _validate_reference_parsing(estates_df, estates_xls_file, "Reference", "Name", "Estates")
        _report_reference_parsing_errors(validation_issues, estates_xls_file, "Estates", output_handler)
        has_critical_errors = len(validation_issues) > 0

    num_estates_added_to_db = 0
    num_blocks_added_to_db = 0

    # Import into DB
    try:
        csr = db_conn.cursor()
        csr.execute("begin")
        for index, row in estates_df.iterrows():
            reference = row["Reference"]
            estate_name = row["Name"]
            # If the property reference begins with a '9' or contains a 'Y' or 'Z',then ignore this data
            if reference is None or reference[0] == "9" or "Y" in reference.upper() or "Z" in reference.upper():
                continue

            # Create property if it doesn't exist, then update it to be an estate
            property_id = get_id_from_ref(csr, "Properties", "property", reference)
            if not property_id:
                # Create the property first
                csr.execute(INSERT_PROPERTY_SQL, (reference,))
                property_id = get_id_from_ref(csr, "Properties", "property", reference)

            if property_id:
                if get_id(csr, SELECT_PROPERTY_ID_FROM_REF_SQL, (reference,)) == property_id:
                    csr.execute(UPDATE_PROPERTY_DETAILS_SQL, (estate_name, property_id))
                    num_estates_added_to_db += 1

                # Add a '00' block for the estate service charges, if not already present
                block_ref = reference + "-00"
                block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
                if not block_id:
                    csr.execute(INSERT_BLOCK_SQL2, (block_ref, estate_name, "P", property_id))
                    num_blocks_added_to_db += 1
        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_estates_added_to_db} estates added to the database.")
        logger.info(f"{num_blocks_added_to_db} estate blocks added to the database.")
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((reference, estate_name)))
        logger.error("No estates or estate blocks have been added to the database")
        csr.execute("rollback")
        raise
    except Exception as ex:
        logger.error(str(ex))
        logger.error("The data which caused the failure is: " + str((reference, estate_name)))
        logger.error("No estates or estate blocks have been added to the database.")
        csr.execute("rollback")
        raise

    # Return whether critical validation errors were found
    return has_critical_errors


def importBankAccounts(db_conn: sqlite3.Connection, bank_accounts_file: str, output_handler: OutputHandler | None = None) -> bool:
    """Import bank accounts data from Excel file into database."""
    # Read Excel spreadsheet into dataframe
    bank_accounts_df = pd.read_excel(bank_accounts_file, "Accounts", dtype={"Reference": str, "Sort Code": str, "Account Number": str})
    bank_accounts_df.replace("nan", "", inplace=True)
    bank_accounts_df.fillna("", inplace=True)

    # Validate for duplicate CL accounts per block and Property/Block designation consistency before inserting any data
    has_critical_errors = False
    if output_handler is not None:
        try:
            _validate_account_uniqueness(bank_accounts_df, bank_accounts_file, output_handler)
        except ValueError:
            # Account uniqueness validation failed - this is critical
            has_critical_errors = True

        # Validate Property/Block designation consistency (doesn't stop import, just reports issues)
        designation_issues = _validate_account_designation_consistency(bank_accounts_df, bank_accounts_file, output_handler)
        if designation_issues:
            logger.warning(f"Found {len(designation_issues)} Property/Block designation issues - see Excel file for details")

    num_bank_accounts_added_to_db = 0

    try:
        csr = db_conn.cursor()
        csr.execute("begin")
        for index, row in bank_accounts_df.iterrows():
            reference = row["Reference"]
            sort_code = row["Sort Code"]
            account_number = row["Account Number"]
            account_type = row["Account Type"]
            property_or_block = row["Property Or Block"]
            client_ref = row["Client Reference"]
            account_name = row["Account Name"]

            # Skip accounts with blank/empty references
            if pd.isna(reference) or str(reference).strip() == "":
                logger.debug(f"Skipping bank account ({sort_code}, {account_number}) with blank reference")
                continue

            property_block = None
            if property_or_block.upper() == "PROPERTY" or property_or_block.upper() == "P":
                property_block = "P"
            elif property_or_block.upper() == "BLOCK" or property_or_block.upper() == "B":
                property_block = "B"
            elif property_or_block == "" or property_or_block is None:
                property_block = ""
            else:
                raise ValueError(f"Unknown property/block type {property_or_block} for bank account ({sort_code}, {account_number})")

            _, block_ref, _ = getPropertyBlockAndTenantRefs(reference)

            block_id = get_id_from_ref(csr, "Blocks", "block", reference)

            _id = get_id(csr, SELECT_BANK_ACCOUNT_SQL1, (sort_code, account_number))
            if sort_code and account_number and not _id:
                csr.execute(
                    INSERT_BANK_ACCOUNT_SQL,
                    (
                        sort_code,
                        account_number,
                        account_type,
                        property_block,
                        client_ref,
                        account_name,
                        block_id,
                    ),
                )
                logger.debug(f"\tAdding bank account ({sort_code}, {account_number}) for property {reference}")
                num_bank_accounts_added_to_db += 1

        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_bank_accounts_added_to_db} bank accounts added to the database.")
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((reference, sort_code, account_number, account_type, property_or_block)))
        logger.error("No bank accounts have been added to the database")
        csr.execute("rollback")
        # Database constraint violations indicate critical errors - don't raise, let caller handle
        has_critical_errors = True
    except Exception as ex:
        logger.error(str(ex))
        logger.error("The data which caused the failure is: " + str((reference, sort_code, account_number, account_type, property_or_block)))
        logger.error("No bank accounts have been added to the database.")
        csr.execute("rollback")
        # Other exceptions indicate critical errors - don't raise, let caller handle
        has_critical_errors = True

    # Return whether critical validation errors were found
    return has_critical_errors


def importIrregularTransactionReferences(db_conn: sqlite3.Connection, anomalous_refs_file: str, output_handler: OutputHandler | None = None) -> bool:
    """Import irregular transaction references from Excel file into database."""
    # Read Excel spreadsheet into dataframe
    anomalous_refs_df = pd.read_excel(anomalous_refs_file, "Sheet1", dtype=str)
    anomalous_refs_df.replace("nan", "", inplace=True)
    anomalous_refs_df.fillna("", inplace=True)

    # Validate tenant reference parsing before import
    if output_handler is not None:
        validation_issues = _validate_reference_parsing(anomalous_refs_df, anomalous_refs_file, "Tenant Reference", "Payment Reference Pattern", "General Idents")
        _report_reference_parsing_errors(validation_issues, anomalous_refs_file, "General Idents", output_handler)

    num_anomalous_refs_added_to_db = 0
    tenant_reference = ""

    try:
        csr = db_conn.cursor()
        csr.execute("begin")
        for index, row in anomalous_refs_df.iterrows():
            tenant_reference = row["Tenant Reference"].strip()
            payment_reference_pattern = row["Payment Reference Pattern"].strip()

            _id = get_id(
                csr,
                SELECT_IRREGULAR_TRANSACTION_REF_ID_SQL,
                (tenant_reference, payment_reference_pattern),
            )
            if tenant_reference and not _id:
                csr.execute(
                    INSERT_IRREGULAR_TRANSACTION_REF_SQL,
                    (tenant_reference, payment_reference_pattern),
                )
                logger.debug(f"\tAdding irregular transaction reference pattern ({tenant_reference}) for tenant {payment_reference_pattern}")
                num_anomalous_refs_added_to_db += 1

        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_anomalous_refs_added_to_db} irregular transaction reference patterns added to the database.")
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("No irregular transaction reference patterns have been added to the database")
        logger.error("The data which caused the failure is: " + str((tenant_reference, payment_reference_pattern)))
        csr.execute("rollback")
        raise
    except Exception as ex:
        logger.error(str(ex))
        logger.error("No irregular transaction reference patterns have been added to the database.")
        logger.error("The data which caused the failure is: " + str((tenant_reference, payment_reference_pattern)))
        csr.execute("rollback")
        raise

    # General idents validation errors don't stop processing
    return False


def importQubeEndOfDayBalancesFile(db_conn: sqlite3.Connection, qube_eod_balances_xls_file: str, output_handler: OutputHandler | None = None) -> bool:
    """Import Qube End of Day balances from Excel file into database."""
    num_charges_added_to_db = 0
    qube_import_errors = []  # Collect import errors for Excel reporting

    # Load and validate spreadsheet
    qube_eod_balances_workbook = load_workbook(qube_eod_balances_xls_file, read_only=True, data_only=True)
    qube_eod_balances_workbook_sheet = qube_eod_balances_workbook.worksheets[0]

    _validate_qube_spreadsheet(qube_eod_balances_workbook_sheet)
    at_date = _extract_qube_date(qube_eod_balances_workbook_sheet)
    qube_eod_balances_df = _prepare_qube_dataframe(qube_eod_balances_xls_file)

    try:
        csr = db_conn.cursor()
        csr.execute("begin")

        # Get type IDs once for efficiency
        type_ids = {
            "auth_creditors": get_id_from_key_table(csr, "type", AUTH_CREDITORS),
            "available_funds": get_id_from_key_table(csr, "type", AVAILABLE_FUNDS),
            "sc_fund": get_id_from_key_table(csr, "type", SC_FUND),
        }

        # Process each row in the dataframe
        found_property = False
        property_ref = block_ref = block_name = None
        block_ref = fund = category = auth_creditors = block_id = None  # For error handling

        for i in range(qube_eod_balances_df.shape[0]):
            property_code_or_fund = qube_eod_balances_df.iloc[i]["PropertyCode / Fund"]
            property_name_or_category = qube_eod_balances_df.iloc[i]["PropertyName / Category"]

            # Check if this is a property/block reference
            try_property_ref, try_block_ref, _ = getPropertyBlockAndTenantRefs(property_code_or_fund)

            if try_property_ref and try_block_ref:
                # Found a new property/block
                found_property = True
                property_ref = try_property_ref
                block_ref = try_block_ref
                block_name = property_name_or_category

            elif found_property and property_code_or_fund in ["Service Charge", "Rent", "Tenant Recharge", "Admin Fund", "Reserve"]:
                # Process fund/category data for current property/block
                fund = property_code_or_fund
                category = property_name_or_category
                auth_creditors = qube_eod_balances_df.iloc[i][AUTH_CREDITORS]
                available_funds = qube_eod_balances_df.iloc[i][AVAILABLE_FUNDS]

                # Skip if we don't have all required data
                if not all([property_ref, block_ref, block_name, fund, category]):
                    logger.warning(f"Missing required data: property_ref={property_ref}, block_ref={block_ref}, block_name={block_name}, fund={fund}, category={category}")
                    continue

                # Type assertions for mypy since we verified they're not None
                assert property_ref is not None
                assert block_ref is not None
                assert block_name is not None

                charges_added, error_info = _process_qube_data(csr, property_code_or_fund, property_ref, block_ref, block_name, fund, category, auth_creditors, available_funds, at_date, type_ids)
                num_charges_added_to_db += charges_added

                # Collect error information for Excel reporting
                if error_info:
                    qube_import_errors.append(error_info)

            elif property_code_or_fund == "Property Totals":
                # Reset for next property
                found_property = False

        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_charges_added_to_db} charges added to the database.")

        # Report any Qube import errors to Excel
        if output_handler is not None:
            _report_qube_import_errors(qube_import_errors, qube_eod_balances_xls_file, output_handler)

    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((block_ref, fund, category, at_date, auth_creditors, block_id)))
        logger.error("No Qube balances have been added to the database.")
        logger.exception(err)
        csr.execute("rollback")
    except Exception as ex:
        logger.error(str(ex))
        logger.error("The data which caused the failure is: " + str((block_ref, fund, category, at_date, auth_creditors, block_id)))
        logger.error("No Qube balances have been added to the database.")
        logger.exception(ex)
        csr.execute("rollback")

    # Return whether critical errors were found (qube import errors are critical)
    return len(qube_import_errors) > 0
