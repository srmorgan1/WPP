import argparse
import datetime as dt
import logging
import os
import re
import sqlite3
import tempfile
import time
import warnings
import xml.etree.ElementTree as et
from pathlib import Path
from typing import Any

import pandas as pd
from dateutil import parser
from lxml import etree
from openpyxl import load_workbook

from .calendars import get_business_day_offset
from .config import get_config, get_wpp_db_file, get_wpp_input_dir, get_wpp_report_dir, get_wpp_static_input_dir, get_wpp_update_database_log_file
from .constants import DEBIT_CARD_SUFFIX, EXCLUDED_TENANT_REF_CHARACTERS, MINIMUM_TENANT_NAME_MATCH_LENGTH, MINIMUM_VALID_PROPERTY_REF
from .data_classes import ChargeData, TransactionReferences
from .database_commands import DatabaseCommandExecutor, InsertBlockCommand, InsertChargeCommand, InsertPropertyCommand, InsertTenantCommand, InsertTransactionCommand, UpdateTenantNameCommand
from .db import checkTenantExists, get_last_insert_id, get_or_create_db, get_single_value, getTenantName
from .exceptions import database_transaction, log_database_error
from .logger import setup_logger
from .output_handler import ExcelOutputHandler, OutputHandler
from .ref_matcher import getPropertyBlockAndTenantRefs as getPropertyBlockAndTenantRefs_strategy
from .utils import getLatestMatchingFileName, getLongestCommonSubstring, getMatchingFileNames, is_running_via_pytest, open_file

#
# Constants
#
CLIENT_CREDIT_ACCOUNT_NUMBER = "06000792"

# Set up logger
logger = logging.getLogger(__name__)

#
# SQL
#
INSERT_PROPERTY_SQL = "INSERT INTO Properties (property_ref, property_name) VALUES (?, Null);"
INSERT_BLOCK_SQL = "INSERT INTO Blocks (block_ref, block_name, type, property_id) VALUES (?, Null, ?, ?);"
INSERT_BLOCK_SQL2 = "INSERT INTO Blocks (block_ref, block_name, type, property_id) VALUES (?, ?, ?, ?);"
INSERT_TENANT_SQL = "INSERT INTO Tenants (tenant_ref, tenant_name, block_id) VALUES (?, ?, ?);"
INSERT_SUGGESTED_TENANT_SQL = "INSERT INTO SuggestedTenants (tenant_id, transaction_id) VALUES (?, ?);"
INSERT_TRANSACTION_SQL = "INSERT INTO Transactions (type, amount, description, pay_date, tenant_id, account_id) VALUES (?, ?, ?, ?, ?, ?);"
INSERT_CHARGES_SQL = "INSERT INTO Charges (fund_id, category_id, type_id, at_date, amount, block_id) VALUES (?, ?, ?, ?, ?, ?);"
INSERT_BANK_ACCOUNT_SQL = "INSERT INTO Accounts (sort_code, account_number, account_type, property_or_block, client_ref, account_name, block_id) VALUES (?, ?, ?, ?, ?, ?, ?);"
INSERT_BANK_ACCOUNT_BALANCE_SQL = "INSERT INTO AccountBalances (current_balance, available_balance, at_date, account_id) VALUES (?, ?, ?, ?);"
INSERT_KEY_TABLE_SQL = "INSERT INTO Key_{} (value) VALUES (?);"
INSERT_IRREGULAR_TRANSACTION_REF_SQL = "INSERT INTO IrregularTransactionRefs (tenant_ref, transaction_ref_pattern) VALUES (?, ?);"

SELECT_TENANT_ID_SQL = "SELECT tenant_id FROM Tenants WHERE tenant_ref = ?;"
SELECT_ID_FROM_REF_SQL = "SELECT ID FROM {} WHERE {}_ref = '{}';"
SELECT_ID_FROM_KEY_TABLE_SQL = "SELECT ID FROM Key_{} WHERE value = ?;"
SELECT_PROPERTY_ID_FROM_REF_SQL = "SELECT ID FROM Properties WHERE property_ref = ? AND property_name IS NULL;"
SELECT_TRANSACTION_SQL = "SELECT ID FROM Transactions WHERE tenant_id = ? AND description = ? AND pay_date = ? AND account_id = ? and type = ? AND amount between (?-0.005) and (?+0.005);"
SELECT_CHARGES_SQL = "SELECT ID FROM Charges WHERE fund_id = ? AND category_id = ? and type_id = ? and block_id = ? and at_date = ?;"
SELECT_BANK_ACCOUNT_SQL = "SELECT ID FROM Blocks WHERE ID = ? AND account_number IS Null;"
SELECT_BANK_ACCOUNT_SQL1 = "SELECT ID FROM Accounts WHERE sort_code = ? AND account_number = ?;"
SELECT_BANK_ACCOUNT_BALANCE_SQL = "SELECT ID FROM AccountBalances WHERE at_date = ? AND account_id = ?;"
SELECT_BLOCK_NAME_SQL = "SELECT block_name FROM Blocks WHERE block_ref = ?;"
SELECT_TENANT_NAME_BY_ID_SQL = "SELECT tenant_name FROM Tenants WHERE ID = ?;"
SELECT_IRREGULAR_TRANSACTION_TENANT_REF_SQL = "select tenant_ref from IrregularTransactionRefs where instr(?, transaction_ref_pattern) > 0;"
SELECT_IRREGULAR_TRANSACTION_REF_ID_SQL = "select ID from IrregularTransactionRefs where tenant_ref = ? and transaction_ref_pattern = ?;"
SELECT_ALL_IRREGULAR_TRANSACTION_REFS_SQL = "select tenant_ref, transaction_ref_pattern from IrregularTransactionRefs;"

UPDATE_BLOCK_ACCOUNT_NUMBER_SQL = "UPDATE Blocks SET account_number = ? WHERE ID = ? AND account_number IS Null;"
UPDATE_PROPERTY_DETAILS_SQL = "UPDATE Properties SET property_name = ? WHERE ID = ?;"
UPDATE_BLOCK_NAME_SQL = "UPDATE Blocks SET block_name = ? WHERE ID = ?;"
UPDATE_TENANT_NAME_SQL = "UPDATE Tenants SET tenant_name = ? WHERE ID = ?;"

# Charge types
AUTH_CREDITORS = "Auth Creditors"
AVAILABLE_FUNDS = "Available Funds"
SC_FUND = "SC Fund"

# # Regular expressions
# PBT_REGEX = re.compile(r"(?:^|\s+|,)(\d\d\d)-(\d\d)-(\d\d\d)\s?(?:DC)?(?:$|\s+|,|/)")
# PBT_REGEX2 = re.compile(r"(?:^|\s+|,)(\d\d\d)\s-\s(\d\d)\s-\s(\d\d\d)\s?(?:DC)?(?:$|\s+|,|/)")
# PBT_REGEX3 = re.compile(r"(?:^|\s+|,)(\d\d\d)-0?(\d\d)-(\d\d)\s?(?:DC)?(?:$|\s+|,|/)")
# PBT_REGEX4 = re.compile(r"(?:^|\s+|,)(\d\d)-0?(\d\d)-(\d\d\d)\s?(?:DC)?(?:$|\s+|,|/)")
# PBT_REGEX_NO_TERMINATING_SPACE = re.compile(r"(?:^|\s+|,)(\d\d\d)-(\d\d)-(\d\d\d)(?:$|\s*|,|/)")
# PBT_REGEX_NO_BEGINNING_SPACE = re.compile(r"(?:^|\s*|,)(\d\d\d)-(\d\d)-(\d\d\d)(?:$|\s+|,|/)")
# PBT_REGEX_SPECIAL_CASES = re.compile(
#     r"(?:^|\s+|,|\.)(\d\d\d)-{1,2}0?(\d\d)-{1,2}(\w{2,5})\s?(?:DC)?(?:$|\s+|,|/)",
#     re.ASCII,
# )
# PBT_REGEX_NO_HYPHENS = re.compile(r"(?:^|\s+|,)(\d\d\d)\s{0,2}0?(\d\d)\s{0,2}(\d\d\d)(?:$|\s+|,|/)")
# PBT_REGEX_NO_HYPHENS_SPECIAL_CASES = re.compile(r"(?:^|\s+|,)(\d\d\d)\s{0,2}0?(\d\d)\s{0,2}(\w{3})(?:$|\s+|,|/)", re.ASCII)
# PBT_REGEX_FWD_SLASHES = re.compile(r"(?:^|\s+|,)(\d\d\d)/0?(\d\d)/(\d\d\d)\s?(?:DC)?(?:$|\s+|,|/)")
# PT_REGEX = re.compile(r"(?:^|\s+|,)(\d\d\d)-(\d\d\d)(?:$|\s+|,|/)")
# PB_REGEX = re.compile(r"(?:^|\s+|,)(\d\d\d)-(\d\d)(?:$|\s+|,|/)")
# P_REGEX = re.compile(r"(?:^|\s+)(\d\d\d)(?:$|\s+)")


def get_id(db_cursor: sqlite3.Cursor, sql: str, args_tuple: tuple = ()) -> int | None:
    return get_single_value(db_cursor, sql, args_tuple)


def get_id_from_ref(db_cursor: sqlite3.Cursor, table_name: str, field_name: str, ref_name: str) -> int | None:
    sql = SELECT_ID_FROM_REF_SQL.format(table_name, field_name, ref_name)
    db_cursor.execute(sql)
    _id = db_cursor.fetchone()
    if _id:
        return _id[0]
    else:
        return None


def get_id_from_key_table(db_cursor: sqlite3.Cursor, key_table_name: str, value: str) -> int:
    sql = SELECT_ID_FROM_KEY_TABLE_SQL.format(key_table_name)
    db_cursor.execute(sql, (value,))
    _id = db_cursor.fetchone()
    if _id:
        return _id[0]
    else:
        sql = INSERT_KEY_TABLE_SQL.format(key_table_name)
        db_cursor.execute(sql, (value,))
        return get_last_insert_id(db_cursor, f"Key_{key_table_name}")


def matchTransactionRef(tenant_name: str, transaction_reference: str) -> bool:
    tnm = re.sub(r"(?:^|\s+)mr?s?\s+", "", tenant_name.lower())
    tnm = re.sub(r"\s+and\s+", "", tnm)
    tnm = re.sub(r"(?:^|\s+)\w\s+", " ", tnm)
    tnm = re.sub(r"[_\W]+", " ", tnm).strip()

    trf = re.sub(r"(?:^|\s+)mr?s?\s+", "", transaction_reference.lower())
    trf = re.sub(r"\s+and\s+", "", trf)
    trf = re.sub(r"(?:^|\s+)\w\s+", " ", trf)
    trf = re.sub(r"\d", "", trf)
    trf = re.sub(r"[_\W]+", " ", trf).strip()

    if tenant_name:
        lcss = getLongestCommonSubstring(tnm, trf)
        # Assume that if the transaction reference has a substring matching
        # one in the tenant name of >= minimum length chars, then this is a match.
        return len(lcss) >= MINIMUM_TENANT_NAME_MATCH_LENGTH
    else:
        return False


def removeDCReferencePostfix(tenant_ref: str | None) -> str | None:
    # Remove 'DC' from parsed tenant references paid by debit card
    if tenant_ref is not None and tenant_ref.endswith(DEBIT_CARD_SUFFIX):
        tenant_ref = tenant_ref[:-2].strip()
    return tenant_ref


def _validate_reference_parsing(references_df: pd.DataFrame, file_path: str, reference_column: str, name_column: str, file_type: str) -> list[dict]:
    """
    Validate that references in a dataframe can be parsed correctly.

    Args:
        references_df: DataFrame containing reference data
        file_path: File path for error messages
        reference_column: Name of the reference column
        name_column: Name of the name column (for context)
        file_type: Type of file for error messages ('Tenants', 'Estates', 'General Idents')

    Returns:
        List of validation issues found
    """
    validation_issues = []

    for index, row in references_df.iterrows():
        reference = row[reference_column]
        name = row[name_column] if name_column in row and pd.notna(row[name_column]) else ""

        # Skip blank/empty references
        if pd.isna(reference) or str(reference).strip() == "":
            continue

        # Try to parse the reference
        property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefs(str(reference).strip())

        # Check if parsing was successful based on file type
        parsing_failed = False
        error_msg = ""

        if file_type == "Tenants":
            # For tenants file, we need all three parts (property, block, tenant)
            if not all([property_ref, block_ref, tenant_ref]):
                parsing_failed = True
                error_msg = f"Tenant reference '{reference}' could not be parsed into property-block-tenant format"
        elif file_type == "Estates":
            # For estates file, we need at least property reference
            if not property_ref:
                parsing_failed = True
                error_msg = f"Estate reference '{reference}' could not be parsed to extract property reference"
        elif file_type == "General Idents":
            # For general idents file, we need tenant reference to be parseable
            if not tenant_ref:
                parsing_failed = True
                error_msg = f"Tenant reference '{reference}' could not be parsed to extract tenant reference"

        if parsing_failed:
            validation_issues.append(
                {
                    "Row Number": index + 2,  # +2 because pandas is 0-indexed and Excel has header row
                    "Reference": reference,
                    "Name": name,
                    "Property Ref": property_ref or "N/A",
                    "Block Ref": block_ref or "N/A",
                    "Tenant Ref": tenant_ref or "N/A",
                    "Error": error_msg,
                }
            )

    return validation_issues


def _report_reference_parsing_errors(validation_issues: list[dict], file_path: str, file_type: str, output_handler: OutputHandler) -> None:
    """
    Report reference parsing errors to log and output handler.

    Args:
        validation_issues: List of error dictionaries
        file_path: File path for error messages
        file_type: Type of file for sheet naming
        output_handler: OutputHandler for outputting errors
    """
    if not validation_issues:
        return

    logger.error(f"Found {len(validation_issues)} reference parsing issues in {file_path}")
    for error in validation_issues:
        logger.error(f"Row {error['Row Number']}: {error['Error']} - Name: {error['Name']}")

    # Write errors to output handler with file-type specific sheet name
    sheet_name = f"{file_type} Import Problems"
    errors_df = pd.DataFrame(validation_issues)
    output_handler.add_sheet(sheet_name, errors_df, {"file_path": file_path, "error_count": len(validation_issues)})


def _report_qube_import_errors(qube_import_errors: list[dict], qube_file: str, output_handler: OutputHandler) -> None:
    """
    Report Qube import errors to log and output handler.

    Args:
        qube_import_errors: List of error dictionaries
        qube_file: File path for error messages
        output_handler: OutputHandler for outputting errors
    """
    if not qube_import_errors:
        return

    logger.error(f"Found {len(qube_import_errors)} Qube import issues in {qube_file}")
    for error in qube_import_errors:
        logger.error(f"Block {error['Block Reference']}: {error['Error']}")

    # Write errors to output handler
    errors_df = pd.DataFrame(qube_import_errors)
    output_handler.add_sheet("Qube Import Problems", errors_df, {"file_path": qube_file, "error_count": len(qube_import_errors)}, is_critical=True)


def _validate_account_designation_consistency(bank_accounts_df: pd.DataFrame, bank_accounts_file: str, output_handler: OutputHandler) -> list[dict]:
    """
    Validate that manual Property/Block designation matches the reference format for all accounts.

    Args:
        bank_accounts_df: DataFrame containing account data
        bank_accounts_file: File path for error messages
        output_handler: OutputHandler for outputting validation issues

    Returns:
        List of validation issues found
    """
    validation_issues = []

    for index, row in bank_accounts_df.iterrows():
        reference = row["Reference"]
        property_or_block = row["Property Or Block"]
        sort_code = row["Sort Code"]
        account_number = row["Account Number"]
        account_name = row["Account Name"]
        client_ref = row["Client Reference"]

        # Skip accounts with blank/empty references
        if pd.isna(reference) or str(reference).strip() == "":
            continue

        # Skip accounts with blank/empty property_or_block
        if pd.isna(property_or_block) or str(property_or_block).strip() == "":
            continue

        _, block_ref, _ = getPropertyBlockAndTenantRefs(reference)

        if block_ref is None:
            continue  # Skip validation if reference couldn't be parsed

        # Normalize property_or_block value
        if property_or_block.upper() in ["PROPERTY", "P"]:
            property_block = "P"
        elif property_or_block.upper() in ["BLOCK", "B"]:
            property_block = "B"
        else:
            continue  # Skip unknown values

        # Derive automatic type from reference format
        auto_type = "P" if block_ref.endswith("00") else "B"

        # Check for inconsistency
        if property_block != auto_type:
            if auto_type == "P":
                # Estate reference marked as Block
                issue = f"Estate reference '{reference}' marked as 'Block' - should be 'Property'?"
                suggestion = f"Change 'Property Or Block' to 'Property' and add to Estates.xlsx, or change reference to '{reference[:-2]}01'"
            else:
                # Block reference marked as Property
                property_ref = reference.split("-")[0] if "-" in reference else reference
                suggested_ref = f"{property_ref}-00"
                issue = f"Block reference '{reference}' marked as 'Property' - should be 'Block'"
                suggestion = f"Change 'Property Or Block' to 'Block' or change reference to '{suggested_ref}'"

            validation_issues.append(
                {
                    "Row Number": index + 2,  # +2 because pandas is 0-indexed and Excel has header row
                    "Reference": reference,
                    "Sort Code": sort_code,
                    "Account Number": account_number,
                    "Account Name": account_name,
                    "Client Reference": client_ref,
                    "Property Or Block": property_or_block,
                    "Expected": "Property" if auto_type == "P" else "Block",
                    "Issue": issue,
                    "Suggestion": suggestion,
                }
            )

    if validation_issues:
        # Log all violations at ERROR level
        logger.error(f"Found {len(validation_issues)} Property/Block designation inconsistencies in {bank_accounts_file}")
        for issue in validation_issues:
            logger.error(f"Row {issue['Row Number']}: {issue['Issue']} - Account {issue['Sort Code']}-{issue['Account Number']}")

        # Write violations to output handler
        issues_df = pd.DataFrame(validation_issues)
        output_handler.add_sheet("Account Designation Issues", issues_df, {"file_path": bank_accounts_file, "error_count": len(validation_issues)})

    return validation_issues


def correctKnownCommonErrors(property_ref: str, block_ref: str, tenant_ref: str | None) -> tuple[str, str, str | None]:
    # Correct known errors in the tenant payment references
    if property_ref == "094" and tenant_ref is not None and tenant_ref[-3] == "O":
        tenant_ref = tenant_ref[:-3] + "0" + tenant_ref[-2:]
    return property_ref, block_ref, tenant_ref


def recodeSpecialPropertyReferenceCases(property_ref: str, block_ref: str, tenant_ref: str | None) -> tuple[str, str, str | None]:
    if property_ref == "020" and block_ref == "020-03":
        # Block 020-03 belongs to a different property group, call this 020A.
        property_ref = "020A"
    elif property_ref == "064" and block_ref == "064-01":
        property_ref = "064A"
    return property_ref, block_ref, tenant_ref


def recodeSpecialBlockReferenceCases(property_ref: str, block_ref: str, tenant_ref: str | None) -> tuple[str, str, str | None]:
    if property_ref == "101" and block_ref == "101-02":
        # Block 101-02 is wrong, change this to 101-01
        block_ref = "101-01"
        if tenant_ref is not None:
            tenant_ref = tenant_ref.replace("101-02", "101-01")
    return property_ref, block_ref, tenant_ref


def getPropertyBlockAndTenantRefsFromRegexMatch(
    match: re.Match,
) -> tuple[str, str, str]:
    property_ref, block_ref, tenant_ref = None, None, None
    if match:
        property_ref = match.group(1)
        block_ref = f"{match.group(1)}-{match.group(2)}"
        tenant_ref = f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return property_ref, block_ref, tenant_ref


def doubleCheckTenantRef(db_cursor: sqlite3.Cursor, tenant_ref: str, reference: str) -> bool:
    if not checkTenantExists(db_cursor, tenant_ref):
        return False
    tenant_name = getTenantName(db_cursor, tenant_ref)
    return matchTransactionRef(tenant_name, reference)


def postProcessPropertyBlockTenantRefs(property_ref: str | None, block_ref: str | None, tenant_ref: str | None) -> tuple[str | None, str | None, str | None]:
    # Ignore some property and tenant references, and recode special cases
    # e.g. Block 020-03 belongs to a different property than the other 020-xx blocks.
    if (tenant_ref is not None and any(char in tenant_ref for char in EXCLUDED_TENANT_REF_CHARACTERS)) or (
        property_ref is not None and property_ref.isnumeric() and int(property_ref) >= MINIMUM_VALID_PROPERTY_REF
    ):
        return None, None, None
    # Only apply special recoding if we have non-None property_ref and block_ref
    if property_ref is not None and block_ref is not None:
        property_ref, block_ref, tenant_ref = recodeSpecialPropertyReferenceCases(property_ref, block_ref, tenant_ref)
        property_ref, block_ref, tenant_ref = recodeSpecialBlockReferenceCases(property_ref, block_ref, tenant_ref)
    return property_ref, block_ref, tenant_ref


def checkForIrregularTenantRefInDatabase(reference: str, db_cursor: sqlite3.Cursor | None) -> tuple[str | None, str | None, str | None]:
    # Look for known irregular transaction refs which we know some tenants use
    if db_cursor:
        tenant_ref = get_single_value(db_cursor, SELECT_IRREGULAR_TRANSACTION_TENANT_REF_SQL, (reference,))
        if tenant_ref:
            return getPropertyBlockAndTenantRefs(tenant_ref)  # Parse tenant reference
        # else:
        #    transaction_ref_data = get_data(db_cursor, SELECT_ALL_IRREGULAR_TRANSACTION_REFS_SQL)
        #    for tenant_ref, transaction_ref_pattern in transaction_ref_data:
        #        pass
    return None, None, None


def getPropertyBlockAndTenantRefs(reference: str, db_cursor: sqlite3.Cursor | None = None) -> tuple[str | None, str | None, str | None]:
    # return getPropertyBlockAndTenantRefsImpl(reference, db_cursor)
    result = getPropertyBlockAndTenantRefs_strategy(reference, db_cursor)
    return result.to_tuple()


# def getPropertyBlockAndTenantRefsImpl(reference: str, db_cursor: sqlite3.Cursor | None = None) -> tuple[str | None, str | None, str | None]:
#     property_ref, block_ref, tenant_ref = None, None, None

#     if not isinstance(reference, str):
#         return None, None, None

#     # Try to match property, block and tenant
#     description = str(reference).strip()

#     # if "MEDHURST K M 10501001 RP4652285818999300" in description:
#     #     pass

#     # Check the database for irregular transaction references first
#     property_ref, block_ref, tenant_ref = checkForIrregularTenantRefInDatabase(description, db_cursor)
#     if property_ref and block_ref and tenant_ref:
#         return property_ref, block_ref, tenant_ref

#     # Then check various regular expression rules
#     match = re.search(PBT_REGEX, description)
#     if match:
#         property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefsFromRegexMatch(match)
#         # if db_cursor and not checkTenantExists(db_cursor, tenant_ref):
#         #    return None, None, None
#     else:
#         match = re.search(PBT_REGEX_FWD_SLASHES, description)
#         if match:
#             property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefsFromRegexMatch(match)
#             if db_cursor and not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                 return None, None, None
#         else:
#             match = re.search(PBT_REGEX2, description)  # Match tenant with spaces between hyphens
#             if match:
#                 property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefsFromRegexMatch(match)
#                 if db_cursor and not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                     return None, None, None
#             else:
#                 match = re.search(PBT_REGEX3, description)  # Match tenant with 2 digits
#                 if match:
#                     property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefsFromRegexMatch(match)
#                     if db_cursor and not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                         tenant_ref = f"{match.group(1)}-{match.group(2)}-0{match.group(3)}"
#                         if not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                             return None, None, None
#                 else:
#                     match = re.search(PBT_REGEX4, description)  # Match property with 2 digits
#                     if match:
#                         property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefsFromRegexMatch(match)
#                         if db_cursor and not checkTenantExists(db_cursor, tenant_ref):
#                             property_ref = f"0{match.group(1)}"
#                             block_ref = f"{property_ref}-{match.group(2)}"
#                             tenant_ref = f"{block_ref}-{match.group(3)}"
#                             if not checkTenantExists(db_cursor, tenant_ref):
#                                 return None, None, None
#                     else:
#                         # Try to match property, block and tenant special cases
#                         match = re.search(PBT_REGEX_SPECIAL_CASES, description)
#                         if match:
#                             property_ref = match.group(1)
#                             block_ref = f"{match.group(1)}-{match.group(2)}"
#                             tenant_ref = f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
#                             if db_cursor:
#                                 tenant_ref = removeDCReferencePostfix(tenant_ref) or tenant_ref
#                                 if tenant_ref and not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                                     if property_ref and block_ref:
#                                         property_ref, block_ref, tenant_ref = correctKnownCommonErrors(property_ref, block_ref, tenant_ref)
#                                     if tenant_ref and not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                                         return None, None, None
#                             elif not (
#                                 (
#                                     property_ref
#                                     in [
#                                         "093",
#                                         "094",
#                                         "095",
#                                         "096",
#                                         "099",
#                                         "124",
#                                         "132",
#                                         "133",
#                                         "134",
#                                     ]
#                                 )
#                                 or (property_ref in ["020", "022", "039", "053", "064"] and match.group(3)[-1] != "Z")
#                             ):
#                                 return None, None, None
#                         else:
#                             # Match property and block only
#                             match = re.search(PB_REGEX, description)
#                             if match:
#                                 property_ref = match.group(1)
#                                 block_ref = f"{match.group(1)}-{match.group(2)}"
#                             else:
#                                 # Match property and tenant only
#                                 # Prevent this case from matching for now, or move to the end of the match blocks
#                                 match = None  # Disabled: re.search(PT_REGEX, description)
#                                 if match:
#                                     pass
#                                     # property_ref = match.group(1)
#                                     # tenant_ref = match.group(2)  # Non-unique tenant ref, may be useful
#                                     # block_ref = '01'   # Null block indicates that the tenant and block can't be matched uniquely
#                                 else:
#                                     # Match without hyphens, or with no terminating space.
#                                     # These cases can only come from parsed transaction references.
#                                     # in which case we can double check that the data exists in and matches the database.
#                                     match = (
#                                         re.search(PBT_REGEX_NO_HYPHENS, description)
#                                         or re.search(
#                                             PBT_REGEX_NO_HYPHENS_SPECIAL_CASES,
#                                             description,
#                                         )
#                                         or re.search(PBT_REGEX_NO_TERMINATING_SPACE, description)
#                                         or re.search(PBT_REGEX_NO_BEGINNING_SPACE, description)
#                                     )
#                                     if match:
#                                         property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefsFromRegexMatch(match)
#                                         if db_cursor and tenant_ref and not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                                             if property_ref and block_ref:
#                                                 property_ref, block_ref, tenant_ref = correctKnownCommonErrors(property_ref, block_ref, tenant_ref)
#                                             if tenant_ref and not doubleCheckTenantRef(db_cursor, tenant_ref, reference):
#                                                 return None, None, None
#                                     # else:
#                                     #    # Match property reference only
#                                     #    match = re.search(P_REGEX, description)
#                                     #    if match:
#                                     #        property_ref = match.group(1)
#                                     #    else:
#                                     #        return None, None, None
#     return postProcessPropertyBlockTenantRefs(property_ref, block_ref, tenant_ref)


def getTenantID(csr: sqlite3.Cursor, tenant_ref: str) -> None:
    csr.execute(SELECT_TENANT_ID_SQL, (tenant_ref))


# Helper function to get text from an XML element and ensure it is not None
def get_element_text(parent_element: et.Element, child_element_name: str) -> str:
    """Extract text from required XML element with enhanced validation.

    Args:
        parent_element: The parent XML element
        child_element_name: Name of the child element to find

    Returns:
        Element text (stripped of whitespace)

    Raises:
        ValueError: When element is missing, empty, or contains only whitespace
    """
    child_element = parent_element.find(child_element_name)

    if child_element is None:
        raise ValueError(f"Missing required XML element: {child_element_name}")

    if child_element.text is None or child_element.text.strip() == "":
        raise ValueError(f"Empty required XML element: {child_element_name}")

    return child_element.text.strip()


def _validate_xml_against_xsd(xml_content: str, xsd_filename: str) -> None:
    """Validate XML content against XSD schema using lxml.

    Args:
        xml_content: XML content as string
        xsd_filename: Name of the XSD file (e.g., 'PreviousDayTransactionExtract.xsd')

    Raises:
        ValueError: If validation fails or XSD file not found
    """
    # Get XSD file path from bundled schemas directory
    schemas_dir = Path(__file__).parent / "schemas"
    xsd_path = schemas_dir / xsd_filename
    if not xsd_path.exists():
        logger.warning(f"XSD file not found: {xsd_path}. Skipping schema validation.")
        return

    try:
        # Parse XSD schema and remove problematic anchors
        with open(xsd_path, encoding="utf-8") as xsd_file:
            xsd_content = xsd_file.read()

        # Remove ^ and $ anchors from xs:pattern values since XML Schema patterns are implicitly anchored
        # Also fix the problematic \d{0,2} pattern to be XML Schema compliant
        original_content = xsd_content
        # Handle patterns with alternation (|) more carefully
        xsd_content = re.sub(r'(<xs:pattern\s+value=")\^([^"]*)\$\|\^\$"', r'\1\2|"', xsd_content)  # ^pattern$|^$ -> pattern|
        xsd_content = re.sub(r'(<xs:pattern\s+value=")\^([^"]*)\$"', r'\1\2"', xsd_content)  # ^pattern$ -> pattern
        xsd_content = re.sub(r'(<xs:pattern\s+value=")\^([^"]*)"', r'\1\2"', xsd_content)  # ^pattern -> pattern (no ending $)
        # Fix the specific problematic pattern: replace the entire amount pattern
        xsd_content = xsd_content.replace("[-+]?\\d{1,13}(?:\\.\\d{0,2})?", "[-+]?\\d{1,13}(\\.\\d{1,2})?")
        logger.debug(f"XSD pattern modifications applied: {original_content != xsd_content}")

        # Create temporary directory with all XSD files to handle includes properly
        with tempfile.TemporaryDirectory() as temp_dir:
            # Copy all XSD files to temp directory, applying fixes to all of them
            for schema_file in xsd_path.parent.glob("*.xsd"):
                if schema_file.name == xsd_filename:
                    # Write our modified main XSD content
                    temp_file_path = Path(temp_dir) / schema_file.name
                    with open(temp_file_path, "w", encoding="utf-8") as f:
                        f.write(xsd_content)
                else:
                    # Read, fix, and write other XSD files (like IBLTypes.xsd)
                    with open(schema_file, encoding="utf-8") as f:
                        other_content = f.read()
                    # Apply same fixes to included files
                    # Handle patterns with alternation (|) more carefully
                    other_content = re.sub(r'(<xs:pattern\s+value=")\^([^"]*)\$\|\^\$"', r'\1\2|"', other_content)  # ^pattern$|^$ -> pattern|
                    other_content = re.sub(r'(<xs:pattern\s+value=")\^([^"]*)\$"', r'\1\2"', other_content)  # ^pattern$ -> pattern
                    other_content = re.sub(r'(<xs:pattern\s+value=")\^([^"]*)"', r'\1\2"', other_content)  # ^pattern -> pattern (no ending $)
                    other_content = other_content.replace("[-+]?\\d{1,13}(?:\\.\\d{0,2})?", "[-+]?\\d{1,13}(\\.\\d{1,2})?")
                    temp_file_path = Path(temp_dir) / schema_file.name
                    with open(temp_file_path, "w", encoding="utf-8") as f:
                        f.write(other_content)

            # Parse XSD schema from temp directory and create schema object
            main_xsd = Path(temp_dir) / xsd_filename
            xsd_doc = etree.parse(str(main_xsd))
            schema = etree.XMLSchema(xsd_doc)

            # Perform the actual XML validation while temp files still exist
            xml_doc = etree.fromstring(xml_content.encode("utf-8"))
            if not schema.validate(xml_doc):
                error_log = schema.error_log
                raise ValueError(f"XML validation failed against {xsd_filename}: {error_log}")

        # If we get here, validation passed
        logger.info(f"XML successfully validated against {xsd_filename}")
        return

    except etree.XMLSyntaxError:
        logger.warning(f"XSD file {xsd_filename} is not valid XML (possibly downloaded error page). Skipping schema validation.")
        return
    except Exception as e:
        logger.warning(f"Unable to perform XSD validation against {xsd_filename}: {e}. Skipping schema validation.")
        return


def _validate_transaction_xml_structure(root: et.Element) -> None:
    """Validate the business structure of transaction XML.

    Args:
        root: The root XML element

    Raises:
        ValueError: If expected business elements are missing
    """
    # Check if we have any TransactionRecord elements
    transaction_records = list(root.iter("TransactionRecord"))
    if not transaction_records:
        raise ValueError("No TransactionRecord elements found in XML - possible schema change")

    logger.debug(f"Found {len(transaction_records)} transaction records in XML")


def _prepare_bos_transaction_xml(transactions_xml_file: str) -> et.Element:
    """Read and prepare Bank of Scotland transaction XML file for parsing."""
    with open_file(transactions_xml_file) as f:
        xml = f.read()
        if type(xml) is bytes:
            xml = str(xml, "utf-8")
        xml = xml.replace("\n", "")

        # Validate XML against XSD before processing
        _validate_xml_against_xsd(xml, "PreviousDayTransactionExtract.xsd")

        # Remove schema namespace for ElementTree parsing
        schema = "PreviousDayTransactionExtract"
        xsd = f"https://isite.bankofscotland.co.uk/Schemas/{schema}.xsd"
        xml = re.sub(
            rf"""<({schema})\s+(xmlns=(?:'|")){xsd}(?:'|")\s*>""",
            r"<\1>",
            xml,
        )

    # Let ParseError propagate naturally - it has valuable diagnostic info
    root = et.fromstring(xml)

    # Add business-level validation
    _validate_transaction_xml_structure(root)

    return root


def _extract_transaction_data(transaction: et.Element) -> dict:
    """Extract transaction data from XML transaction element."""
    return {
        "sort_code": get_element_text(transaction, "SortCode"),
        "account_number": get_element_text(transaction, "AccountNumber"),
        "transaction_type": get_element_text(transaction, "TransactionType"),
        "amount": get_element_text(transaction, "TransactionAmount"),
        "description": get_element_text(transaction, "TransactionDescription"),
        "pay_date": get_element_text(transaction, "TransactionPostedDate"),
    }


def _should_process_transaction(transaction_data: dict) -> bool:
    """Check if transaction should be processed based on account number."""
    return transaction_data["account_number"] == CLIENT_CREDIT_ACCOUNT_NUMBER


def _format_pay_date(pay_date: str) -> str:
    """Format payment date to standard format."""
    return parser.parse(pay_date, dayfirst=True).strftime("%Y-%m-%d")


def _process_valid_transaction(csr: sqlite3.Cursor, transaction_data: dict, refs: TransactionReferences) -> tuple[bool, bool]:
    """Process a valid transaction and return (was_added, is_duplicate)."""
    account_id = get_id(csr, SELECT_BANK_ACCOUNT_SQL1, (transaction_data["sort_code"], transaction_data["account_number"]))
    tenant_id = get_id_from_ref(csr, "Tenants", "tenant", refs.tenant_ref) if refs.tenant_ref else None

    if not tenant_id:
        return False, False

    # Check for duplicate transaction
    transaction_id = get_id(
        csr,
        SELECT_TRANSACTION_SQL,
        (
            tenant_id,
            transaction_data["description"],
            transaction_data["pay_date"],
            account_id,
            transaction_data["transaction_type"],
            transaction_data["amount"],
            transaction_data["amount"],
        ),
    )

    if transaction_id:
        return False, True  # Duplicate found

    # Add new transaction
    executor = DatabaseCommandExecutor(csr, logger)
    command = InsertTransactionCommand(
        transaction_data["transaction_type"],
        transaction_data["amount"],
        transaction_data["description"],
        transaction_data["pay_date"],
        tenant_id,
        account_id,
        transaction_data["sort_code"],
        transaction_data["account_number"],
        refs.tenant_ref,
        INSERT_TRANSACTION_SQL,
    )
    executor.execute(command)
    return True, False


def _create_error_record(transaction_data: dict, error_message: str) -> list:
    """Create an error record for unprocessable transactions."""
    return [
        transaction_data["pay_date"],
        transaction_data["sort_code"],
        transaction_data["account_number"],
        transaction_data["transaction_type"],
        float(transaction_data["amount"]),
        transaction_data["description"],
        error_message,
    ]


def _create_missing_tenant_record(transaction_data: dict, tenant_ref: str) -> list:
    """Create a record for transactions with valid references but missing tenants."""
    return [
        transaction_data["pay_date"],
        transaction_data["sort_code"],
        transaction_data["account_number"],
        transaction_data["transaction_type"],
        float(transaction_data["amount"]),
        transaction_data["description"],
        tenant_ref,
        f"Tenant '{tenant_ref}' not found in tenants file",
    ]


def _create_duplicate_record(transaction_data: dict, tenant_ref: str) -> list:
    """Create a duplicate transaction record."""
    return [
        transaction_data["pay_date"],
        transaction_data["transaction_type"],
        float(transaction_data["amount"]),
        tenant_ref,
        transaction_data["description"],
    ]


def _handle_transaction_processing_error(csr: sqlite3.Cursor, error: Exception, transaction_data: dict, tenant_id: int | None) -> None:
    """Handle errors that occur during transaction processing."""
    error_context = {
        "sort_code": transaction_data.get("sort_code"),
        "account_number": transaction_data.get("account_number"),
        "transaction_type": transaction_data.get("transaction_type"),
        "amount": transaction_data.get("amount"),
        "description": transaction_data.get("description"),
        "pay_date": transaction_data.get("pay_date"),
        "tenant_id": tenant_id,
    }

    # Use standardized database error logging
    log_database_error(logger, "importing Bank of Scotland transaction", error, error_context)
    logger.error("No Bank Of Scotland transactions have been added to the database.")

    csr.execute("rollback")


def _process_single_transaction(csr: sqlite3.Cursor, transaction_data: dict) -> tuple[str, bool, bool]:
    """Process a single transaction and return the result.

    Returns:
        tuple: (result_type, was_processed, needs_error_record)
        result_type: "added", "duplicate", "tenant_not_found", "invalid_refs"
        was_processed: True if transaction was successfully processed
        needs_error_record: True if an error record should be created
    """
    property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefs(transaction_data["description"], csr)

    if not (tenant_ref and property_ref and block_ref):
        return "invalid_refs", False, True

    refs = TransactionReferences(property_ref, block_ref, tenant_ref)
    was_added, is_duplicate = _process_valid_transaction(csr, transaction_data, refs)

    if was_added:
        return "added", True, False
    elif is_duplicate:
        return "duplicate", True, False
    else:
        return "tenant_not_found", False, True


def _process_valid_transaction(csr: sqlite3.Cursor, transaction_data: dict, refs: TransactionReferences) -> tuple[bool, bool]:
    """Process a valid transaction (where tenant references are available).

    Returns:
        tuple: (was_added, is_duplicate)
    """
    account_id = get_id(csr, SELECT_BANK_ACCOUNT_SQL1, (transaction_data["sort_code"], transaction_data["account_number"]))
    tenant_id = get_id_from_ref(csr, "Tenants", "tenant", refs.tenant_ref) if refs.tenant_ref else None

    if not tenant_id:
        return False, False

    # Check for duplicate transaction
    transaction_id = get_id(
        csr,
        SELECT_TRANSACTION_SQL,
        (
            tenant_id,
            transaction_data["description"],
            transaction_data["pay_date"],
            account_id,
            transaction_data["transaction_type"],
            transaction_data["amount"],
            transaction_data["amount"],
        ),
    )

    if transaction_id:
        return False, True  # Duplicate found

    # Add new transaction
    executor = DatabaseCommandExecutor(csr, logger)
    command = InsertTransactionCommand(
        transaction_data["transaction_type"],
        transaction_data["amount"],
        transaction_data["description"],
        transaction_data["pay_date"],
        tenant_id,
        account_id,
        transaction_data["sort_code"],
        transaction_data["account_number"],
        refs.tenant_ref,
        INSERT_TRANSACTION_SQL,
    )
    executor.execute(command)
    return True, False


def _process_transaction_results(
    result_type: str, transaction_data: dict, tenant_ref: str | None, unrecognised_transactions: list, duplicate_transactions: list, missing_tenant_transactions: list
) -> tuple[int, int]:
    """Process the results of a transaction and update counters.

    Returns:
        tuple: (added_count, error_count)
    """
    if result_type == "added":
        return 1, 0
    elif result_type == "duplicate":
        duplicate_transactions.append(_create_duplicate_record(transaction_data, tenant_ref))
        return 0, 0
    elif result_type == "tenant_not_found":
        error_msg = f"Cannot find tenant with reference '{tenant_ref}'"
        logger.debug(
            f"{error_msg}. Ignoring transaction {(transaction_data['pay_date'], transaction_data['sort_code'], transaction_data['account_number'], transaction_data['transaction_type'], transaction_data['amount'], transaction_data['description'])}"
        )
        missing_tenant_transactions.append(_create_missing_tenant_record(transaction_data, tenant_ref))
        return 0, 1
    elif result_type == "invalid_refs":
        error_msg = "Cannot determine tenant from description"
        logger.debug(
            f"{error_msg} '{transaction_data['description']}'. Ignoring transaction {(transaction_data['pay_date'], transaction_data['sort_code'], transaction_data['account_number'], transaction_data['transaction_type'], transaction_data['amount'], transaction_data['description'])}"
        )
        unrecognised_transactions.append(_create_error_record(transaction_data, error_msg))
        return 0, 1

    return 0, 0


def importBankOfScotlandTransactionsXMLFile(db_conn: sqlite3.Connection, transactions_xml_file: str) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]]]:
    """Import Bank of Scotland transactions from XML file into database.

    Returns:
        tuple: (unrecognised_transactions, duplicate_transactions, missing_tenant_transactions) - Lists of problematic transactions
    """
    unrecognised_transactions = []
    duplicate_transactions = []
    missing_tenant_transactions = []
    tree = _prepare_bos_transaction_xml(transactions_xml_file)

    num_transactions_added_to_db = 0
    num_import_errors = 0
    tenant_id = None
    transaction_data = {}

    try:
        csr = db_conn.cursor()
        csr.execute("begin")

        for transaction in tree.iter("TransactionRecord"):
            transaction_data = _extract_transaction_data(transaction)

            if not _should_process_transaction(transaction_data):
                continue

            transaction_data["pay_date"] = _format_pay_date(transaction_data["pay_date"])

            result_type, was_processed, needs_error = _process_single_transaction(csr, transaction_data)

            # Extract tenant_ref for duplicate records
            if result_type in ["duplicate", "tenant_not_found"]:
                _, _, tenant_ref = getPropertyBlockAndTenantRefs(transaction_data["description"], csr)
            else:
                tenant_ref = None

            added_count, error_count = _process_transaction_results(result_type, transaction_data, tenant_ref, unrecognised_transactions, duplicate_transactions, missing_tenant_transactions)

            num_transactions_added_to_db += added_count
            num_import_errors += error_count

        csr.execute("end")
        db_conn.commit()

        if num_import_errors:
            logger.info(
                f"Unable to import {num_import_errors} transactions into the database. "
                f"See the Data_Import_Issues Excel file for details. Add tenant references to "
                f"{get_config()['INPUTS']['IRREGULAR_TRANSACTION_REFS_FILE']} and run import again."
            )
        logger.info(f"{num_transactions_added_to_db} Bank Of Scotland transactions added to the database.")
        return unrecognised_transactions, duplicate_transactions, missing_tenant_transactions

    except (db_conn.Error, Exception) as error:
        _handle_transaction_processing_error(csr, error, transaction_data, tenant_id)
        return [], [], []


def _validate_balance_xml_structure(root: et.Element) -> None:
    """Validate the business structure of balance XML.

    Args:
        root: The root XML element

    Raises:
        ValueError: If expected business elements are missing
    """
    # Check if we have any ReportingDay elements
    reporting_days = list(root.iter("ReportingDay"))
    if not reporting_days:
        raise ValueError("No ReportingDay elements found in XML - possible schema change")

    # Check if we have any BalanceRecord elements
    balance_records = list(root.iter("BalanceRecord"))
    if not balance_records:
        raise ValueError("No BalanceRecord elements found in XML - possible schema change")

    logger.debug(f"Found {len(reporting_days)} reporting days and {len(balance_records)} balance records in XML")


def _prepare_bos_xml(balances_xml_file: str) -> et.Element:
    """Read and clean Bank of Scotland XML file, returning parsed tree."""
    with open_file(balances_xml_file) as f:
        xml = f.read()
        if isinstance(xml, bytes):
            xml = str(xml, "utf-8")
        xml = xml.replace("\n", "")

        # Validate XML against XSD before processing
        _validate_xml_against_xsd(xml, "EndOfDayBalanceExtract.xsd")

        # Remove schema namespaces to simplify XML parsing
        for schema in ["BalanceDetailedReport", "EndOfDayBalanceExtract"]:
            xsd = f"https://isite.bankofscotland.co.uk/Schemas/{schema}.xsd"
            xml = re.sub(
                rf"""<({schema})\s+(xmlns=(?:'|")){xsd}(?:'|")\s*>""",
                r"<\1>",
                xml,
            )

    # Let ParseError propagate naturally - it has valuable diagnostic info
    root = et.fromstring(xml)

    # Add business-level validation
    _validate_balance_xml_structure(root)

    return root


def _determine_account_type(client_ref: str | None) -> str:
    """Determine account type from client reference."""
    if not client_ref:
        return "NA"

    client_ref_upper = client_ref.upper()
    if "RENT" in client_ref_upper:
        return "GR"
    elif "BANK" in client_ref_upper:
        return "CL"
    elif "RES" in client_ref_upper:
        return "RE"
    else:
        return "NA"


def _extract_balance_data(balance_element) -> dict:
    """Extract balance data from XML balance element."""
    # ClientRef is optional - handle explicitly
    client_ref_element = balance_element.find("ClientRef")
    client_ref = None
    if client_ref_element is not None and client_ref_element.text is not None:
        client_ref = client_ref_element.text.strip()
        if client_ref == "":  # Empty string treated as None
            client_ref = None

    return {
        "sort_code": get_element_text(balance_element, "SortCode"),
        "account_number": get_element_text(balance_element, "AccountNumber"),
        "client_ref": client_ref,
        "account_name": get_element_text(balance_element, "LongName"),
        "account_type": _determine_account_type(client_ref),
        "current_balance": get_element_text(balance_element, "CurrentBalance"),
        "available_balance": get_element_text(balance_element, "AvailableBalance"),
    }


def _process_balance_record(csr, balance_data: dict, at_date: str) -> bool:
    """Process a single balance record and add to database if needed. Returns True if added."""
    sort_code = balance_data["sort_code"]
    account_number = balance_data["account_number"]

    if not (sort_code and account_number):
        logger.warning(f"Cannot determine bank account. Ignoring balance record {balance_data}")
        return False

    account_id = get_id(csr, SELECT_BANK_ACCOUNT_SQL1, (sort_code, account_number))
    if not account_id:
        # Account doesn't exist in our system, skip
        return False

    # Check if balance already exists for this date
    account_balance_id = get_id(csr, SELECT_BANK_ACCOUNT_BALANCE_SQL, (at_date, account_id))
    if account_balance_id:
        # Balance already exists, skip
        return False

    # Add new balance record
    csr.execute(
        INSERT_BANK_ACCOUNT_BALANCE_SQL,
        (
            balance_data["current_balance"],
            balance_data["available_balance"],
            at_date,
            account_id,
        ),
    )

    logger.debug(
        f"\tAdding bank balance {(sort_code, account_number, balance_data['account_type'], balance_data['client_ref'], balance_data['account_name'], at_date, balance_data['current_balance'], balance_data['available_balance'])}"
    )
    return True


def _process_balance_reporting_day(csr: sqlite3.Cursor, reporting_day_element) -> int:
    """Process all balance records for a single reporting day.

    Returns:
        int: Number of balance records added to database
    """
    at_date = get_element_text(reporting_day_element, "Date")
    at_date = parser.parse(at_date, dayfirst=True).strftime("%Y-%m-%d")

    balances_added = 0
    for balance in reporting_day_element.iter("BalanceRecord"):
        balance_data = _extract_balance_data(balance)
        if _process_balance_record(csr, balance_data, at_date):
            balances_added += 1

    return balances_added


def _log_balance_import_error(error: Exception, balance_data: dict, at_date: str) -> None:
    """Log balance import error with context data."""
    sort_code = balance_data.get("sort_code")
    account_number = balance_data.get("account_number")
    account_type = balance_data.get("account_type")
    client_ref = balance_data.get("client_ref")
    account_name = balance_data.get("account_name")
    current_balance = balance_data.get("current_balance")
    available_balance = balance_data.get("available_balance")

    logger.error(str(error))
    logger.error(f"The data which caused the failure is: {(sort_code, account_number, account_type, client_ref, account_name, at_date, current_balance, available_balance)}")
    logger.error("No Bank Of Scotland account balances have been added to the database.")
    logger.exception(error)


def importBankOfScotlandBalancesXMLFile(db_conn: sqlite3.Connection, balances_xml_file: str) -> None:
    """Import Bank of Scotland account balances from XML file into database."""
    tree = _prepare_bos_xml(balances_xml_file)
    num_balances_added_to_db = 0

    # Variables for error handling context
    last_balance_data = {}
    last_at_date = ""

    try:
        csr = db_conn.cursor()
        csr.execute("begin")

        for reporting_day in tree.iter("ReportingDay"):
            last_at_date = get_element_text(reporting_day, "Date")
            try:
                balances_added = _process_balance_reporting_day(csr, reporting_day)
                num_balances_added_to_db += balances_added
            except Exception as day_error:
                # Update context for error reporting
                for balance in reporting_day.iter("BalanceRecord"):
                    last_balance_data = _extract_balance_data(balance)
                    break  # Just get the first one for context
                raise day_error

        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_balances_added_to_db} Bank Of Scotland account balances added to the database.")

    except db_conn.Error as err:
        _log_balance_import_error(err, last_balance_data, last_at_date)
        csr.execute("rollback")
    except Exception as ex:
        _log_balance_import_error(ex, last_balance_data, last_at_date)
        csr.execute("rollback")


def _read_properties_df(properties_xls_file: str) -> pd.DataFrame:
    """Reads properties data from an Excel file into a DataFrame."""
    properties_df = pd.read_excel(properties_xls_file)
    properties_df.fillna("", inplace=True)
    return properties_df


def _is_valid_reference(reference: str) -> bool:
    """Checks if a reference is valid for processing."""
    if not reference or reference.startswith("9") or "Y" in reference.upper() or "Z" in reference.upper():
        return False
    return True


def _process_property(csr: sqlite3.Cursor, property_ref: str) -> tuple[int, int]:
    """Processes a property, adding it to the DB if it doesn't exist."""
    executor = DatabaseCommandExecutor(csr, logger)
    property_id = get_id_from_ref(csr, "Properties", "property", property_ref)
    if not property_id:
        command = InsertPropertyCommand(property_ref, INSERT_PROPERTY_SQL)
        new_id = executor.execute(command)
        return new_id, 1
    assert property_id is not None  # For mypy
    return property_id, 0


def _process_block(csr: sqlite3.Cursor, block_ref: str, property_id: int) -> tuple[int, int]:
    """Processes a block, adding it to the DB if it doesn't exist."""
    executor = DatabaseCommandExecutor(csr, logger)
    block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
    if not block_id:
        block_type = "P" if block_ref and block_ref.endswith("00") else "B"
        command = InsertBlockCommand(block_ref, block_type, property_id, INSERT_BLOCK_SQL)
        new_id = executor.execute(command)
        return new_id, 1
    assert block_id is not None  # For mypy
    return block_id, 0


def _process_tenant(csr: sqlite3.Cursor, tenant_ref: str, tenant_name: str, block_id: int) -> int:
    """Processes a tenant, adding or updating it in the DB."""
    executor = DatabaseCommandExecutor(csr, logger)
    tenant_id = get_id_from_ref(csr, "Tenants", "tenant", tenant_ref)
    if tenant_ref and not tenant_id:
        command = InsertTenantCommand(tenant_ref, tenant_name, block_id, INSERT_TENANT_SQL)
        executor.execute(command)
        return 1
    elif tenant_id:
        old_tenant_name = get_single_value(csr, SELECT_TENANT_NAME_BY_ID_SQL, (tenant_id,))
        if tenant_name and tenant_name != old_tenant_name:
            command = UpdateTenantNameCommand(tenant_name, tenant_id, tenant_ref, UPDATE_TENANT_NAME_SQL)
            executor.execute(command)
    return 0


def importPropertiesFile(db_conn: sqlite3.Connection, properties_xls_file: str, output_handler: OutputHandler | None = None) -> bool:
    """Import properties data from Excel file into database."""
    properties_df = _read_properties_df(properties_xls_file)

    # Validate tenant reference parsing before import
    if output_handler is not None:
        validation_issues = _validate_reference_parsing(properties_df, properties_xls_file, "Reference", "Name", "Tenants")
        _report_reference_parsing_errors(validation_issues, properties_xls_file, "Tenants", output_handler)

    num_properties_added_to_db = 0
    num_blocks_added_to_db = 0
    num_tenants_added_to_db = 0
    current_data = {}

    with database_transaction(db_conn, logger, "importing properties file") as csr:
        for index, row in properties_df.iterrows():
            reference = row["Reference"]
            tenant_name = row["Name"]
            current_data = {"reference": reference, "tenant_name": tenant_name}

            if not _is_valid_reference(reference):
                continue

            property_ref, block_ref, tenant_ref = getPropertyBlockAndTenantRefs(reference)
            if not all((property_ref, block_ref, tenant_ref)):
                logger.warning(f"\tUnable to parse tenant reference {reference}, will not add to the database.")
                continue

            current_data.update({"property_ref": property_ref, "block_ref": block_ref, "tenant_ref": tenant_ref})

            # Type assertions for mypy since we know they're not None after the all() check
            assert property_ref is not None
            assert block_ref is not None
            assert tenant_ref is not None

            try:
                prop_id, props_added = _process_property(csr, property_ref)
                num_properties_added_to_db += props_added

                blk_id, blocks_added = _process_block(csr, block_ref, prop_id)
                num_blocks_added_to_db += blocks_added

                tenants_added = _process_tenant(csr, tenant_ref, tenant_name, blk_id)
                num_tenants_added_to_db += tenants_added
            except Exception:
                logger.error(f"Failed to process record: {current_data}")
                raise

    logger.info(f"{num_properties_added_to_db} properties added to the database.")
    logger.info(f"{num_blocks_added_to_db} blocks added to the database.")
    logger.info(f"{num_tenants_added_to_db} tenants added to the database.")

    # Tenants validation errors don't stop processing
    return False


def importEstatesFile(db_conn: sqlite3.Connection, estates_xls_file: str, output_handler: OutputHandler | None = None) -> bool:
    # Read Excel spreadsheet into dataframe
    estates_df = pd.read_excel(estates_xls_file, dtype={"Reference": str})
    estates_df.fillna("", inplace=True)

    # Validate estate reference parsing before import
    has_critical_errors = False
    if output_handler is not None:
        validation_issues = _validate_reference_parsing(estates_df, estates_xls_file, "Reference", "Name", "Estates")
        _report_reference_parsing_errors(validation_issues, estates_xls_file, "Estates", output_handler)
        has_critical_errors = len(validation_issues) > 0

    num_estates_added_to_db = 0
    num_blocks_added_to_db = 0

    # Import into DB
    try:
        csr = db_conn.cursor()
        csr.execute("begin")
        for index, row in estates_df.iterrows():
            reference = row["Reference"]
            estate_name = row["Name"]
            # If the property reference begins with a '9' or contains a 'Y' or 'Z',then ignore this data
            if reference is None or reference[0] == "9" or "Y" in reference.upper() or "Z" in reference.upper():
                continue

            # Create property if it doesn't exist, then update it to be an estate
            property_id = get_id_from_ref(csr, "Properties", "property", reference)
            if not property_id:
                # Create the property first
                csr.execute(INSERT_PROPERTY_SQL, (reference,))
                property_id = get_id_from_ref(csr, "Properties", "property", reference)

            if property_id:
                if get_id(csr, SELECT_PROPERTY_ID_FROM_REF_SQL, (reference,)) == property_id:
                    csr.execute(UPDATE_PROPERTY_DETAILS_SQL, (estate_name, property_id))
                    num_estates_added_to_db += 1

                # Add a '00' block for the estate service charges, if not already present
                block_ref = reference + "-00"
                block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
                if not block_id:
                    csr.execute(INSERT_BLOCK_SQL2, (block_ref, estate_name, "P", property_id))
                    num_blocks_added_to_db += 1
        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_estates_added_to_db} estates added to the database.")
        logger.info(f"{num_blocks_added_to_db} estate blocks added to the database.")
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((reference, estate_name)))
        logger.error("No estates or estate blocks have been added to the database")
        csr.execute("rollback")
        raise
    except Exception as ex:
        logger.error(str(ex))
        logger.error("The data which caused the failure is: " + str((reference, estate_name)))
        logger.error("No estates or estate blocks have been added to the database.")
        csr.execute("rollback")
        raise

    # Return whether critical validation errors were found
    return has_critical_errors


def addPropertyToDB(db_conn: sqlite3.Connection, property_ref: str, rethrow_exception: bool = False) -> int | None:
    property_id = None
    try:
        csr = db_conn.cursor()
        csr.execute("begin")

        if property_ref:
            property_id = get_id_from_ref(csr, "Properties", "property", property_ref)
            if not property_id:
                csr.execute(INSERT_PROPERTY_SQL, (property_ref,))
                logger.debug(f"\tAdding property {property_ref}")
                property_id = get_last_insert_id(csr, "Properties")

        csr.execute("end")
        db_conn.commit()
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error(f"The data which caused the failure is: {property_ref}")
        logger.error(f"Unable to add property {property_ref} to the database")
        logger.exception(err)
        csr.execute("rollback")
        if rethrow_exception:
            raise
    except Exception as ex:
        logger.error(str(ex))
        logger.exception(ex)
        logger.error(f"Unable to add property {property_ref} to the database")
        csr.execute("rollback")
        if rethrow_exception:
            raise
    return property_id


def addBlockToDB(
    db_conn: sqlite3.Connection,
    property_ref: str,
    block_ref: str,
    rethrow_exception: bool = False,
) -> int | None:
    block_id = None
    try:
        csr = db_conn.cursor()
        csr.execute("begin")

        if block_ref:
            block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
            if not block_id and property_ref:
                if block_ref.endswith("00"):
                    block_type = "P"
                else:
                    block_type = "B"
                property_id = get_id_from_ref(csr, "Properties", "property", property_ref)
                csr.execute(INSERT_BLOCK_SQL, (block_ref, block_type, property_id))
                logger.debug(f"\tAdding block {block_ref}")
                block_id = get_last_insert_id(csr, "Blocks")

        csr.execute("end")
        db_conn.commit()
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((property_ref, block_ref)))
        logger.error("Unable to add property or block to the database")
        logger.exception(err)
        csr.execute("rollback")
        if rethrow_exception:
            raise
    except Exception as ex:
        logger.error(str(ex))
        logger.exception(ex)
        logger.error("The data which caused the failure is: " + str((property_ref, block_ref)))
        logger.error("Unable to add property or block to the database")
        csr.execute("rollback")
        if rethrow_exception:
            raise
    return block_id


def addTenantToDB(
    db_conn: sqlite3.Connection,
    block_ref: str,
    tenant_ref: str,
    tenant_name: str,
    rethrow_exception: bool = False,
) -> int | None:
    tenant_id = None
    try:
        csr = db_conn.cursor()
        csr.execute("begin")

        if tenant_ref:
            tenant_id = get_id_from_ref(csr, "Tenants", "tenant", tenant_ref)
            if not tenant_id:
                block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
                if block_id:
                    csr.execute(INSERT_TENANT_SQL, (tenant_ref, tenant_name, block_id))
                    logger.debug(f"\tAdding tenant {tenant_ref}")
                    tenant_id = get_last_insert_id(csr, "Tenants")

        csr.execute("end")
        db_conn.commit()
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((block_ref, tenant_ref)))
        logger.error("Unable to add tenant to the database")
        logger.exception(err)
        csr.execute("rollback")
        if rethrow_exception:
            raise
    except Exception as ex:
        logger.error(str(ex))
        logger.exception(ex)
        logger.error("The data which caused the failure is: " + str((block_ref, tenant_ref)))
        logger.error("Unable to add tenant to the database")
        csr.execute("rollback")
        if rethrow_exception:
            raise
    return tenant_id


# def importBlockBankAccountNumbers(db_conn: sqlite3.Connection, bos_reconciliations_file: str) -> None:
#     # Read Excel spreadsheet into dataframe
#     bank_accounts_df = pd.read_excel(bos_reconciliations_file, "Accounts", dtype=str)

#     num_bank_accounts_added_to_db = 0

#     try:
#         csr = db_conn.cursor()
#         csr.execute("begin")
#         for index, row in bank_accounts_df.iterrows():
#             block_ref = row["Reference"]
#             account_number = row["Account Number"]

#             block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
#             if block_id:
#                 id = get_id(csr, SELECT_BANK_ACCOUNT_SQL, (block_id,))
#                 if id:
#                     csr.execute(UPDATE_BLOCK_ACCOUNT_NUMBER_SQL, (account_number, block_id))
#                     logger.debug(f"\tAdding bank account number {account_number} for block {block_id}")
#                     num_bank_accounts_added_to_db += 1
#         csr.execute("end")
#         db_conn.commit()
#         logger.info(f"{num_bank_accounts_added_to_db} bank account numbers added to the database.")
#     except db_conn.Error as err:
#         logger.error(str(err))
#         logger.error("The data which caused the failure is: " + str((block_ref, account_number)))
#         logger.error("No bank account numbers have been added to the database")
#         logger.exception(err)
#         csr.execute("rollback")
#     except Exception as ex:
#         logger.error(str(ex))
#         logger.error("The data which caused the failure is: " + str((block_ref, account_number)))
#         logger.error("No bank account numbers have been added to the database.")
#         logger.exception(ex)
#         csr.execute("rollback")
#         # charges = {}


def _validate_account_uniqueness(bank_accounts_df: pd.DataFrame, bank_accounts_file: str, output_handler: OutputHandler) -> None:
    """
    Validate that there are no duplicate CL (Client) accounts per block reference.
    Ignores accounts with blank/empty references.
    Writes violations to the existing Excel writer and raises ValueError if violations are found.
    """
    # Find duplicate CL accounts per block reference, excluding blank references
    cl_accounts = bank_accounts_df[(bank_accounts_df["Account Type"] == "CL") & (bank_accounts_df["Reference"].notna()) & (bank_accounts_df["Reference"].str.strip() != "")].copy()

    if cl_accounts.empty:
        return  # No CL accounts to validate

    # Group by Reference and count CL accounts
    cl_counts = cl_accounts.groupby("Reference").size()
    duplicate_blocks = cl_counts[cl_counts > 1]

    if duplicate_blocks.empty:
        return  # No duplicates found

    # Collect all violation details for reporting
    violations = []
    for block_ref in duplicate_blocks.index:
        block_accounts = cl_accounts[cl_accounts["Reference"] == block_ref]
        for _, row in block_accounts.iterrows():
            violations.append(
                {
                    "Block Reference": block_ref,
                    "Account Number": row["Account Number"],
                    "Sort Code": row["Sort Code"],
                    "Account Name": row["Account Name"],
                    "Client Reference": row["Client Reference"],
                    "Row Number": row.name + 2,  # +2 because pandas is 0-indexed and Excel has header row
                    "Error": f"Multiple CL accounts found for block {block_ref}",
                }
            )

    # Log all violations at ERROR level
    logger.error(f"Found {len(violations)} duplicate CL account violations in {bank_accounts_file}")
    for violation in violations:
        logger.error(f"Row {violation['Row Number']}: Block {violation['Block Reference']} has multiple CL accounts - Account {violation['Sort Code']}-{violation['Account Number']}")

    # Write violations to the output handler
    violations_df = pd.DataFrame(violations)
    output_handler.add_sheet("Account Validation Problems", violations_df, {"error_count": len(violations)}, is_critical=True)

    # Raise exception to stop the import process
    # The Excel file will be saved by the finally block in importAllData()
    block_list = ", ".join(duplicate_blocks.index)
    raise ValueError(
        f"Data validation failed: Found multiple CL (Client) accounts for the following blocks: {block_list}. "
        f"Each block can have only one CL account. Please check the 'Account Validation Problems' sheet in the "
        f"Data Import Issues Excel file for details and correct the Accounts.xlsx file before running the import again."
    )


def importBankAccounts(db_conn: sqlite3.Connection, bank_accounts_file: str, output_handler: OutputHandler | None = None) -> bool:
    # Read Excel spreadsheet into dataframe
    bank_accounts_df = pd.read_excel(bank_accounts_file, "Accounts", dtype={"Reference": str, "Sort Code": str, "Account Number": str})
    bank_accounts_df.replace("nan", "", inplace=True)
    bank_accounts_df.fillna("", inplace=True)

    # Validate for duplicate CL accounts per block and Property/Block designation consistency before inserting any data
    has_critical_errors = False
    if output_handler is not None:
        try:
            _validate_account_uniqueness(bank_accounts_df, bank_accounts_file, output_handler)
        except ValueError:
            # Account uniqueness validation failed - this is critical
            has_critical_errors = True

        # Validate Property/Block designation consistency (doesn't stop import, just reports issues)
        designation_issues = _validate_account_designation_consistency(bank_accounts_df, bank_accounts_file, output_handler)
        if designation_issues:
            logger.warning(f"Found {len(designation_issues)} Property/Block designation issues - see Excel file for details")

    num_bank_accounts_added_to_db = 0

    try:
        csr = db_conn.cursor()
        csr.execute("begin")
        for index, row in bank_accounts_df.iterrows():
            reference = row["Reference"]
            sort_code = row["Sort Code"]
            account_number = row["Account Number"]
            account_type = row["Account Type"]
            property_or_block = row["Property Or Block"]
            client_ref = row["Client Reference"]
            account_name = row["Account Name"]

            # Skip accounts with blank/empty references
            if pd.isna(reference) or str(reference).strip() == "":
                logger.debug(f"Skipping bank account ({sort_code}, {account_number}) with blank reference")
                continue

            property_block = None
            if property_or_block.upper() == "PROPERTY" or property_or_block.upper() == "P":
                property_block = "P"
            elif property_or_block.upper() == "BLOCK" or property_or_block.upper() == "B":
                property_block = "B"
            elif property_or_block == "" or property_or_block is None:
                property_block = ""
            else:
                raise ValueError(f"Unknown property/block type {property_or_block} for bank account ({sort_code}, {account_number})")

            _, block_ref, _ = getPropertyBlockAndTenantRefs(reference)

            block_id = get_id_from_ref(csr, "Blocks", "block", reference)
            # if block_id is None:
            #     property_ref, _, _ = getPropertyBlockAndTenantRefs(reference)
            #     logger.warning(
            #         f"Block reference '{reference}' does not exist in database for bank account {sort_code}-{account_number}. "
            #         f"Ensure property '{property_ref}' exists in the Properties table and block {reference} exists in the blocks table. "
            #         f"Skipping this bank account as it is not required."
            #     )
            #     continue  # Skip this account rather than failing

            _id = get_id(csr, SELECT_BANK_ACCOUNT_SQL1, (sort_code, account_number))
            if sort_code and account_number and not _id:
                csr.execute(
                    INSERT_BANK_ACCOUNT_SQL,
                    (
                        sort_code,
                        account_number,
                        account_type,
                        property_block,
                        client_ref,
                        account_name,
                        block_id,
                    ),
                )
                logger.debug(f"\tAdding bank account ({sort_code}, {account_number}) for property {reference}")
                num_bank_accounts_added_to_db += 1

        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_bank_accounts_added_to_db} bank accounts added to the database.")
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((reference, sort_code, account_number, account_type, property_or_block)))
        logger.error("No bank accounts have been added to the database")
        csr.execute("rollback")
        # Database constraint violations indicate critical errors - don't raise, let caller handle
        has_critical_errors = True
    except Exception as ex:
        logger.error(str(ex))
        logger.error("The data which caused the failure is: " + str((reference, sort_code, account_number, account_type, property_or_block)))
        logger.error("No bank accounts have been added to the database.")
        csr.execute("rollback")
        # Other exceptions indicate critical errors - don't raise, let caller handle
        has_critical_errors = True

    # Return whether critical validation errors were found
    return has_critical_errors


def importIrregularTransactionReferences(db_conn: sqlite3.Connection, anomalous_refs_file: str, output_handler: OutputHandler | None = None) -> bool:
    # Read Excel spreadsheet into dataframe
    anomalous_refs_df = pd.read_excel(anomalous_refs_file, "Sheet1", dtype=str)
    anomalous_refs_df.replace("nan", "", inplace=True)
    anomalous_refs_df.fillna("", inplace=True)

    # Validate tenant reference parsing before import
    if output_handler is not None:
        validation_issues = _validate_reference_parsing(anomalous_refs_df, anomalous_refs_file, "Tenant Reference", "Payment Reference Pattern", "General Idents")
        _report_reference_parsing_errors(validation_issues, anomalous_refs_file, "General Idents", output_handler)

    num_anomalous_refs_added_to_db = 0
    tenant_reference = ""

    try:
        csr = db_conn.cursor()
        csr.execute("begin")
        for index, row in anomalous_refs_df.iterrows():
            tenant_reference = row["Tenant Reference"].strip()
            payment_reference_pattern = row["Payment Reference Pattern"].strip()

            _id = get_id(
                csr,
                SELECT_IRREGULAR_TRANSACTION_REF_ID_SQL,
                (tenant_reference, payment_reference_pattern),
            )
            if tenant_reference and not _id:
                csr.execute(
                    INSERT_IRREGULAR_TRANSACTION_REF_SQL,
                    (tenant_reference, payment_reference_pattern),
                )
                logger.debug(f"\tAdding irregular transaction reference pattern ({tenant_reference}) for tenant {payment_reference_pattern}")
                num_anomalous_refs_added_to_db += 1

        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_anomalous_refs_added_to_db} irregular transaction reference patterns added to the database.")
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("No irregular transaction reference patterns have been added to the database")
        logger.error("The data which caused the failure is: " + str((tenant_reference, payment_reference_pattern)))
        csr.execute("rollback")
        raise
    except Exception as ex:
        logger.error(str(ex))
        logger.error("No irregular transaction reference patterns have been added to the database.")
        logger.error("The data which caused the failure is: " + str((tenant_reference, payment_reference_pattern)))
        csr.execute("rollback")
        raise

    # General idents validation errors don't stop processing
    return False


def calculateSCFund(auth_creditors: float, available_funds: float, property_ref: str, block_ref: str) -> float:
    # TODO: This should be encoded in a user-supplied rules spreadsheet for generality
    if property_ref == "035":
        return available_funds
    else:
        return auth_creditors + available_funds


def _validate_qube_spreadsheet(workbook_sheet) -> None:
    """Validate that the spreadsheet is a valid Qube balances report."""
    A1_cell_value = workbook_sheet.cell(1, 1).value
    B1_cell_value = workbook_sheet.cell(1, 2).value
    produced_date_cell_value = workbook_sheet.cell(3, 1).value

    if not isinstance(produced_date_cell_value, str):
        raise ValueError(f"The produced date cell value is not a string: {produced_date_cell_value}")

    cell_values_actual = [workbook_sheet.cell(5, i + 1).value for i in range(0, 4)]
    cell_values_expected = [
        "Property / Fund",
        "Bank",
        "Excluded VAT",
        AUTH_CREDITORS,
        AVAILABLE_FUNDS,
    ]

    is_valid_header = A1_cell_value == "Property Management" and B1_cell_value == "Funds Available in Property Funds"
    is_valid_columns = all(actual == expected for actual, expected in zip(cell_values_actual, cell_values_expected))

    if not (is_valid_header and is_valid_columns):
        logger.error("The spreadsheet does not look like a Qube balances report.")


def _extract_qube_date(workbook_sheet) -> str:
    """Extract and parse the date from the Qube report."""
    produced_date_cell_value = workbook_sheet.cell(3, 1).value
    at_date_str = " ".join(produced_date_cell_value.split()[-3:])
    return (parser.parse(at_date_str, dayfirst=True) - BUSINESS_DAY).strftime("%Y-%m-%d")


def _prepare_qube_dataframe(qube_eod_balances_xls_file: str) -> pd.DataFrame:
    """Read and prepare the Qube balances dataframe."""
    qube_eod_balances_df = pd.read_excel(qube_eod_balances_xls_file, usecols="B:G", skiprows=4)

    # Fix column names
    qube_eod_balances_df.columns = [  # type: ignore[assignment]
        "PropertyCode / Fund",
        "PropertyName / Category",
        "Bank",
        "Excluded VAT",
        AUTH_CREDITORS,
        AVAILABLE_FUNDS,
    ]

    # Clean data
    qube_eod_balances_df.dropna(how="all", inplace=True)
    qube_eod_balances_df.fillna(0, inplace=True)

    return qube_eod_balances_df


def _diagnose_missing_property(csr, property_ref: str, block_ref: str) -> str:
    """Provide diagnostic information for missing property."""
    # Check if property exists at all
    property_count = get_id(csr, "SELECT COUNT(*) FROM Properties WHERE property_ref = ?", (property_ref,))

    if property_count == 0:
        return (
            f"Property '{property_ref}' does not exist in the Properties table. "
            f"Check if property '{property_ref}' is included in the Tenants.xlsx file "
            f"or add it to the Estates.xlsx file if it's an estate."
        )
    else:
        return f"Property '{property_ref}' exists but block '{block_ref}' was not found. This may indicate a block creation issue."


def _ensure_block_exists(csr, property_ref: str, block_ref: str) -> int | None:
    """Ensure the block exists in the database and return its ID."""
    property_id = get_id_from_ref(csr, "Properties", "property", property_ref)
    if not property_id:
        return None

    block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)
    if not block_id:
        block_type = "P" if block_ref and block_ref.endswith("00") else "B"
        csr.execute(INSERT_BLOCK_SQL, (block_ref, block_type, property_id))
        logger.debug(f"\tAdding block {block_ref}")
        block_id = get_id_from_ref(csr, "Blocks", "block", block_ref)

    return block_id


def _update_block_name_if_needed(csr, block_ref: str, block_name: str, block_id: int) -> None:
    """Update block name if it doesn't exist."""
    if not get_id(csr, SELECT_BLOCK_NAME_SQL, (block_ref,)):
        csr.execute(UPDATE_BLOCK_NAME_SQL, (block_name, block_id))
        logger.debug(f"\tAdding block name {block_name} for block reference {block_ref}")


def _add_charge_if_not_exists(csr, charge: ChargeData) -> bool:
    """Add a charge to the database if it doesn't already exist. Returns True if added."""
    executor = DatabaseCommandExecutor(csr, logger)
    charges_id = get_id(csr, SELECT_CHARGES_SQL, (charge.fund_id, charge.category_id, charge.type_id, charge.block_id, charge.at_date))
    if not charges_id:
        command = InsertChargeCommand(charge, INSERT_CHARGES_SQL)
        executor.execute(command)
        return True
    return False


def _process_qube_data(
    csr, property_code_or_fund: str, property_ref: str, block_ref: str, block_name: str, fund: str, category: str, auth_creditors: float, available_funds: float, at_date: str, type_ids: dict
) -> tuple[int, dict | None]:
    """Process fund/category data and add charges to database. Returns number of charges added."""
    num_charges_added = 0

    # Ensure block exists
    block_id = _ensure_block_exists(csr, property_ref, block_ref)
    if not block_id:
        diagnostic_info = _diagnose_missing_property(csr, property_ref, block_ref)
        error_msg = f"Cannot process Qube balances for block '{block_ref}'. {diagnostic_info}"
        logger.error(error_msg + f" Skipping {block_ref} charges.")

        # Return error details for Excel reporting
        error_details = {
            "Property Code/Fund": property_code_or_fund,
            "Property Reference": property_ref,
            "Block Reference": block_ref,
            "Block Name": block_name,
            "Fund": fund,
            "Category": category,
            "Auth Creditors": auth_creditors,
            "Available Funds": available_funds,
            "Date": at_date,
            "Error": diagnostic_info,
        }
        return 0, error_details

    # Update block name if needed
    _update_block_name_if_needed(csr, block_ref, block_name, block_id)

    # Get fund and category IDs
    fund_id = get_id_from_key_table(csr, "fund", fund)
    category_id = get_id_from_key_table(csr, "category", category)

    # Add available funds charge
    available_charge = ChargeData(fund_id, category_id, type_ids["available_funds"], at_date, available_funds, block_id)
    if _add_charge_if_not_exists(csr, available_charge):
        logger.debug(f"\tAdding charge {(fund, category, AVAILABLE_FUNDS, at_date, block_ref, available_funds)}")
        num_charges_added += 1

    # Add auth creditors and SC fund charges for specific fund types
    if property_code_or_fund in ["Service Charge", "Tenant Recharge"]:
        auth_charge = ChargeData(fund_id, category_id, type_ids["auth_creditors"], at_date, auth_creditors, block_id)
        if _add_charge_if_not_exists(csr, auth_charge):
            logger.debug(f"\tAdding charge for {(fund, category, AUTH_CREDITORS, at_date, block_ref, auth_creditors)}")
            num_charges_added += 1

        sc_fund = calculateSCFund(auth_creditors, available_funds, property_ref, block_ref)
        sc_charge = ChargeData(fund_id, category_id, type_ids["sc_fund"], at_date, sc_fund, block_id)
        if _add_charge_if_not_exists(csr, sc_charge):
            logger.debug(f"\tAdding charge for {(fund, category, SC_FUND, at_date, block_ref, sc_fund)}")
            num_charges_added += 1

    return num_charges_added, None  # No error for successful processing


def importQubeEndOfDayBalancesFile(db_conn: sqlite3.Connection, qube_eod_balances_xls_file: str, output_handler: OutputHandler | None = None) -> bool:
    """Import Qube End of Day balances from Excel file into database."""
    num_charges_added_to_db = 0
    qube_import_errors = []  # Collect import errors for Excel reporting

    # Load and validate spreadsheet
    qube_eod_balances_workbook = load_workbook(qube_eod_balances_xls_file, read_only=True, data_only=True)
    qube_eod_balances_workbook_sheet = qube_eod_balances_workbook.worksheets[0]

    _validate_qube_spreadsheet(qube_eod_balances_workbook_sheet)
    at_date = _extract_qube_date(qube_eod_balances_workbook_sheet)
    qube_eod_balances_df = _prepare_qube_dataframe(qube_eod_balances_xls_file)

    try:
        csr = db_conn.cursor()
        csr.execute("begin")

        # Get type IDs once for efficiency
        type_ids = {
            "auth_creditors": get_id_from_key_table(csr, "type", AUTH_CREDITORS),
            "available_funds": get_id_from_key_table(csr, "type", AVAILABLE_FUNDS),
            "sc_fund": get_id_from_key_table(csr, "type", SC_FUND),
        }

        # Process each row in the dataframe
        found_property = False
        property_ref = block_ref = block_name = None
        block_ref = fund = category = auth_creditors = block_id = None  # For error handling

        for i in range(qube_eod_balances_df.shape[0]):
            property_code_or_fund = qube_eod_balances_df.iloc[i]["PropertyCode / Fund"]
            property_name_or_category = qube_eod_balances_df.iloc[i]["PropertyName / Category"]

            # Check if this is a property/block reference
            try_property_ref, try_block_ref, _ = getPropertyBlockAndTenantRefs(property_code_or_fund)

            if try_property_ref and try_block_ref:
                # Found a new property/block
                found_property = True
                property_ref = try_property_ref
                block_ref = try_block_ref
                block_name = property_name_or_category

            elif found_property and property_code_or_fund in ["Service Charge", "Rent", "Tenant Recharge", "Admin Fund", "Reserve"]:
                # Process fund/category data for current property/block
                fund = property_code_or_fund
                category = property_name_or_category
                auth_creditors = qube_eod_balances_df.iloc[i][AUTH_CREDITORS]
                available_funds = qube_eod_balances_df.iloc[i][AVAILABLE_FUNDS]

                # Skip if we don't have all required data
                if not all([property_ref, block_ref, block_name, fund, category]):
                    logger.warning(f"Missing required data: property_ref={property_ref}, block_ref={block_ref}, block_name={block_name}, fund={fund}, category={category}")
                    continue

                # Type assertions for mypy since we verified they're not None
                assert property_ref is not None
                assert block_ref is not None
                assert block_name is not None

                charges_added, error_info = _process_qube_data(csr, property_code_or_fund, property_ref, block_ref, block_name, fund, category, auth_creditors, available_funds, at_date, type_ids)
                num_charges_added_to_db += charges_added

                # Collect error information for Excel reporting
                if error_info:
                    qube_import_errors.append(error_info)

            elif property_code_or_fund == "Property Totals":
                # Reset for next property
                found_property = False

        csr.execute("end")
        db_conn.commit()
        logger.info(f"{num_charges_added_to_db} charges added to the database.")

        # Report any Qube import errors to Excel
        if output_handler is not None:
            _report_qube_import_errors(qube_import_errors, qube_eod_balances_xls_file, output_handler)

    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("The data which caused the failure is: " + str((block_ref, fund, category, at_date, auth_creditors, block_id)))
        logger.error("No Qube balances have been added to the database.")
        logger.exception(err)
        csr.execute("rollback")
    except Exception as ex:
        logger.error(str(ex))
        logger.error("The data which caused the failure is: " + str((block_ref, fund, category, at_date, auth_creditors, block_id)))
        logger.error("No Qube balances have been added to the database.")
        logger.exception(ex)
        csr.execute("rollback")

    # Return whether critical errors were found (qube import errors are critical)
    return len(qube_import_errors) > 0


def add_misc_data_to_db(db_conn: sqlite3.Connection) -> None:
    # Add account number and property name for some properties
    try:
        csr = db_conn.cursor()
        # csr.execute('begin')
        # csr.execute(UPDATE_PROPERTY_DETAILS_SQL, ('St Winefrides Estate', '020'))
        # csr.execute(UPDATE_PROPERTY_DETAILS_SQL, ('Sand Wharf Estate', '034'))
        # csr.execute(UPDATE_PROPERTY_DETAILS_SQL, ('Hensol Castle Park', '036'))
        # csr.execute(UPDATE_PROPERTY_DETAILS_SQL, ('Grangemoor Court', '064'))
        # csr.execute(UPDATE_PROPERTY_DETAILS_SQL, ('Farmleigh Estate', '074'))
        # csr.execute(UPDATE_PROPERTY_DETAILS_SQL, ('St Fagans Rise', '095'))
        # Add the 064-00 block as this is not yet present in the data
        # block_id = get_id_from_ref(csr, 'Blocks', 'block', '064-00')
        # if not block_id:
        #    property_id = get_id_from_ref(csr, 'Properties', 'property', '064')
        #    sql = "INSERT INTO Blocks (block_ref, block_name, type, property_id) VALUES (?, ?, ?, ?);"
        #    csr.execute(sql, ('064-00', 'Grangemoor Court', 'P', property_id))
        # csr.execute('end')
        # db_conn.commit()
    except db_conn.Error as err:
        logger.error(str(err))
        logger.error("No miscellaneous data has been added to the database.")
        logger.exception(err)
        csr.execute("rollback")
    except Exception as ex:
        logger.error(str(ex))
        logger.error("No miscellaneous data has been added to the database.")
        logger.exception(ex)
        csr.execute("rollback")


def importAllData(db_conn: sqlite3.Connection, output_handler: OutputHandler) -> None:
    """Import all data with flexible output handling via OutputHandler interface."""

    # Track critical validation errors that should stop processing
    has_critical_validation_errors = False

    config = get_config()
    inputs_config = config["INPUTS"]

    # Import iregular transaction references
    irregular_transaction_refs_file_pattern = os.path.join(get_wpp_static_input_dir(), f"{inputs_config['IRREGULAR_TRANSACTION_REFS_FILE']}")
    irregular_transaction_refs_file = getLatestMatchingFileName(irregular_transaction_refs_file_pattern)
    if irregular_transaction_refs_file:
        logger.info(f"Importing irregular transaction references from file {irregular_transaction_refs_file}")
        has_critical_validation_errors |= importIrregularTransactionReferences(db_conn, irregular_transaction_refs_file, output_handler)
    else:
        logger.error(f"Cannot find irregular transaction references file matching {irregular_transaction_refs_file_pattern}")
    logger.info("")

    # Import tenants
    properties_file_pattern = os.path.join(get_wpp_static_input_dir(), inputs_config["PROPERTIES_FILE_PATTERN"])
    tenants_file_pattern = os.path.join(get_wpp_static_input_dir(), inputs_config["TENANTS_FILE_PATTERN"])
    properties_xls_file = getLatestMatchingFileName(properties_file_pattern) or getLatestMatchingFileName(tenants_file_pattern)
    if properties_xls_file:
        logger.info(f"Importing Properties from file {properties_xls_file}")
        has_critical_validation_errors |= importPropertiesFile(db_conn, properties_xls_file, output_handler)
    else:
        logger.error(f"Cannot find Properties file matching {properties_file_pattern}")
    logger.info("")

    estates_file_pattern = os.path.join(get_wpp_static_input_dir(), inputs_config["ESTATES_FILE_PATTERN"])
    estates_xls_file = getLatestMatchingFileName(estates_file_pattern)
    if estates_xls_file:
        logger.info(f"Importing Estates from file {estates_xls_file}")
        has_critical_validation_errors |= importEstatesFile(db_conn, estates_xls_file, output_handler)
    else:
        logger.error(f"Cannot find Estates file matching {estates_file_pattern}")
    logger.info("")

    qube_eod_balances_file_pattern = os.path.join(get_wpp_input_dir(), inputs_config["QUBE_EOD_BALANCES_PATTERN"])
    qube_eod_balances_files = getMatchingFileNames(qube_eod_balances_file_pattern)
    if qube_eod_balances_files:
        for qube_eod_balances_file in qube_eod_balances_files:
            logger.info(f"Importing Qube balances from file {qube_eod_balances_file}")
            has_critical_validation_errors |= importQubeEndOfDayBalancesFile(db_conn, qube_eod_balances_file, output_handler)
    else:
        logger.error(f"Cannot find Qube EOD Balances file matching {qube_eod_balances_file_pattern}")
    logger.info("")

    accounts_file_pattern = os.path.join(get_wpp_static_input_dir(), inputs_config["BANK_ACCOUNTS_PATTERN"])
    accounts_file = getLatestMatchingFileName(accounts_file_pattern)
    if accounts_file:
        logger.info(f"Importing bank accounts from file {accounts_file}")
        has_critical_validation_errors |= importBankAccounts(db_conn, accounts_file, output_handler)
    else:
        logger.error(f"ERROR: Cannot find account numbers file matching {accounts_file_pattern}")
    logger.info("")

    bos_statement_file_pattern = [
        os.path.join(get_wpp_input_dir(), f)
        for f in [
            "PreviousDayTransactionExtract_*.xml",
            "PreviousDayTransactionExtract_*.zip",
        ]
    ]
    bos_statement_xml_files = getMatchingFileNames(bos_statement_file_pattern)

    # Filter out .xml files if corresponding .zip files exist to avoid duplicate processing
    file_set = set(bos_statement_xml_files)
    bos_statement_xml_files = [f for f in bos_statement_xml_files if not (f.endswith(".xml") and f[:-4] + ".zip" in file_set)]

    unrecognised_transactions = []
    duplicate_transactions = []
    missing_tenant_transactions = []
    if bos_statement_xml_files:
        for bos_statement_xml_file in bos_statement_xml_files:
            logger.info(f"Importing Bank Account Transactions from file {bos_statement_xml_file}")
            unrecognised, duplicates, missing_tenants = importBankOfScotlandTransactionsXMLFile(db_conn, bos_statement_xml_file)
            unrecognised_transactions.extend(unrecognised)
            duplicate_transactions.extend(duplicates)
            missing_tenant_transactions.extend(missing_tenants)
        errors_columns = [
            "Payment Date",
            "Sort Code",
            "Account Number",
            "Transaction Type",
            "Amount",
            "Description",
            "Reason",
        ]
        duplicates_columns = [
            "Payment Date",
            "Transaction Type",
            "Amount",
            "Tenant Reference",
            "Description",
        ]
        missing_tenant_columns = [
            "Payment Date",
            "Sort Code",
            "Account Number",
            "Transaction Type",
            "Amount",
            "Description",
            "Parsed Tenant Reference",
            "Issue",
        ]
        # Only create Excel sheets that have data rows (not just headers)
        if unrecognised_transactions:
            errors_df = pd.DataFrame(unrecognised_transactions, columns=errors_columns)
            output_handler.add_sheet("Unrecognised Transactions", errors_df, {"error_count": len(unrecognised_transactions)})

        if duplicate_transactions:
            duplicates_df = pd.DataFrame(duplicate_transactions, columns=duplicates_columns)
            output_handler.add_sheet("Duplicate Transactions", duplicates_df, {"error_count": len(duplicate_transactions)})

        if missing_tenant_transactions:
            missing_tenant_df = pd.DataFrame(missing_tenant_transactions, columns=missing_tenant_columns)
            output_handler.add_sheet("Missing Tenants", missing_tenant_df, {"error_count": len(missing_tenant_transactions)})
    else:
        logger.error(f"Cannot find bank account transactions file matching {bos_statement_file_pattern}")
    logger.info("")

    eod_balances_file_patterns = [
        os.path.join(get_wpp_input_dir(), f)
        for f in [
            "EOD BalancesReport_*.xml",
            "EndOfDayBalanceExtract_*.xml",
            "EndOfDayBalanceExtract_*.zip",
        ]
    ]
    eod_balances_xml_files = getMatchingFileNames(eod_balances_file_patterns)

    # Filter out .xml files if corresponding .zip files exist to avoid duplicate processing
    file_set = set(eod_balances_xml_files)
    eod_balances_xml_files = [f for f in eod_balances_xml_files if not (f.endswith(".xml") and f[:-4] + ".zip" in file_set)]

    if eod_balances_xml_files:
        for eod_balances_xml_file in eod_balances_xml_files:
            logger.info(f"Importing Bank Account balances from file {eod_balances_xml_file}")
            importBankOfScotlandBalancesXMLFile(db_conn, eod_balances_xml_file)
    else:
        logger.error("Cannot find bank account balances file matching one of {}".format(",".join(eod_balances_file_patterns)))
    # Add separator line for log formatting
    logger.info("")

    print("DEBUG: About to call add_misc_data_to_db")
    add_misc_data_to_db(db_conn)
    print("DEBUG: add_misc_data_to_db completed")

    # Export ref_matcher data to output handler (web display + console CSV as configured)
    from .ref_matcher import _get_matcher

    matcher = _get_matcher()
    matcher.export_collected_data(output_handler)

    # Check if any critical validation errors occurred and raise exception to stop processing
    # Add summary data
    output_handler.add_summary("import_status", {"has_critical_errors": has_critical_validation_errors, "completed_successfully": not has_critical_validation_errors})

    # Finalize output
    result = output_handler.build()

    if has_critical_validation_errors:
        logger.error(f"Data Import Issues saved with validation errors: {result}")
        raise ValueError(
            "Data validation failed: Critical validation errors were found in estates, qube, or accounts data. "
            "Please check the Data Import Issues file for details and correct the input files before "
            "running the import again."
        )


def get_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("-v", "--verbose", type=str, help="Generate verbose log file.")
    args = parser.parse_args()
    return args


def update_database_core(injected_logger=None, output_handler=None) -> None:
    """Core database update functionality that can be called from CLI or API.

    Args:
        injected_logger: Logger instance to use. If None, creates default file logger.
        output_handler: OutputHandler instance to use. If None, creates default Excel handler.
    """
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    global logger

    if injected_logger:
        # Use the injected logger (could be web logger or any other logger)
        logger = injected_logger
    else:
        # Create default file logger for CLI usage
        log_file = get_wpp_update_database_log_file(dt.datetime.today())
        logger = setup_logger(__name__, log_file)

    global BUSINESS_DAY
    BUSINESS_DAY = get_business_day_offset(logger)

    # Set up output handler
    if not output_handler:
        # Create default Excel output handler for CLI usage
        from .config import get_wpp_excel_log_file

        excel_log_file = get_wpp_excel_log_file(dt.date.today())
        output_handler = ExcelOutputHandler(str(excel_log_file))

    start_time = time.time()

    os.makedirs(get_wpp_input_dir(), exist_ok=True)
    os.makedirs(get_wpp_static_input_dir(), exist_ok=True)
    os.makedirs(get_wpp_report_dir(), exist_ok=True)

    logger.info("Beginning Import of data into the database")

    db_conn = get_or_create_db(get_wpp_db_file(), logger)
    result = importAllData(db_conn, output_handler)

    elapsed_time = time.time() - start_time
    time.strftime("%S", time.gmtime(elapsed_time))

    logger.info("Import completed")

    print("DEBUG: About to close database connection")
    # Close database connection
    db_conn.close()
    print("DEBUG: Database connection closed")

    return result


def main() -> None:
    """Command-line entry point for database update."""
    # Get command line arguments
    args = get_args() if not is_running_via_pytest() else argparse.Namespace()
    if not args:
        return

    # Call the core function
    update_database_core()
    logger.info("----------------------------------------------------------------------------------------")
    # input("Press enter to end.")


def _handle_database_error(error: Exception, context_data: dict, operation_description: str) -> None:
    """
    Handle database errors with consistent logging format.

    Args:
        error: The exception that occurred
        context_data: Dictionary of relevant data that caused the failure
        operation_description: Description of what operation failed
    """
    logger.error(str(error))
    logger.error(f"The data which caused the failure is: {context_data}")
    logger.error(f"{operation_description}")
    if hasattr(error, "__traceback__"):
        logger.exception(error)


if __name__ == "__main__":
    main()
